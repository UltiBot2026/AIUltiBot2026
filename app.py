"""
Ultiphoton Solar Power OPC - Advanced AI Chatbot for Facebook Messenger
Features: Language Detection, Conversation Memory, Typing Indicator, Quick Replies, 
Analytics, Auto-Greeting, Business Hours Response
"""

from flask import Flask, request
import requests
import json
import os
import sys
import time
import re
import hmac
import hashlib
from datetime import datetime
from collections import defaultdict
import sqlite3
from threading import Lock

app = Flask(__name__)

# Configuration
PAGE_ACCESS_TOKEN = os.getenv("PAGE_ACCESS_TOKEN", "").strip()
if not PAGE_ACCESS_TOKEN:
    PAGE_ACCESS_TOKEN = "default_token_here"

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "").strip()
GOOGLE_MAPS_API_KEY = os.getenv("GOOGLE_MAPS_API_KEY", "").strip()
APP_SECRET = os.getenv("APP_SECRET", "").strip()          # Meta App Secret for webhook signature verification
ANALYTICS_TOKEN = os.getenv("ANALYTICS_TOKEN", "").strip()  # Bearer token to protect /analytics endpoint
PAGE_ID = "516699488185698"
VERIFY_TOKEN = "ultiphoton_solar_verify_2026"

# Business Hours (Philippine Time - PST/PHT)
BUSINESS_HOURS = {
    "start": 8,      # 8 AM
    "end": 18,       # 6 PM
    "days": [0, 1, 2, 3, 4, 5, 6]  # All days (0=Monday, 6=Sunday)
}

# Database lock for thread safety
db_lock = Lock()

print("\n" + "="*70)
print("🤖 ULTIPHOTON SOLAR POWER OPC - ADVANCED AI CHATBOT")
print("="*70)
print(f"✅ Page ID: {PAGE_ID}")
print(f"✅ Access Token: {'✓ SET' if PAGE_ACCESS_TOKEN else '✗ NOT SET'}")
print(f"✅ OpenAI Key: {'✓ SET' if OPENAI_API_KEY else '✗ NOT SET'}")
print("✅ Features: Language Detection | Conversation Memory | Typing Indicator")
print("✅ Features: Quick Replies | Analytics | Auto-Greeting | Business Hours")
print("="*70 + "\n")
sys.stdout.flush()

# Initialize Database
def init_database():
    """Initialize SQLite database for conversation history and analytics"""
    try:
        with db_lock:
            conn = sqlite3.connect('/tmp/ultiphoton_chatbot.db')
            cursor = conn.cursor()
            
            # Conversation history table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS conversations (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id TEXT NOT NULL,
                    message TEXT NOT NULL,
                    response TEXT NOT NULL,
                    language TEXT,
                    timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                    faq_matched BOOLEAN DEFAULT 0
                )
            ''')
            
            # Analytics table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS analytics (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id TEXT NOT NULL,
                    faq_key TEXT,
                    keyword TEXT,
                    timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # User preferences table
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS user_preferences (
                    user_id TEXT PRIMARY KEY,
                    language TEXT DEFAULT 'auto',
                    first_message_sent BOOLEAN DEFAULT 0,
                    last_greeting_date TEXT DEFAULT NULL,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            # Migration: add last_greeting_date column if it doesn't exist yet
            try:
                cursor.execute('ALTER TABLE user_preferences ADD COLUMN last_greeting_date TEXT DEFAULT NULL')
            except Exception:
                pass  # Column already exists
            # Migration: add pending_panel_qty for wattage clarification flow
            try:
                cursor.execute('ALTER TABLE user_preferences ADD COLUMN pending_panel_qty INTEGER DEFAULT NULL')
            except Exception:
                pass  # Column already exists
            # Migration: add pending_wattage for installer price flow
            try:
                cursor.execute('ALTER TABLE user_preferences ADD COLUMN pending_wattage INTEGER DEFAULT NULL')
            except Exception:
                pass  # Column already exists
            
            conn.commit()
            conn.close()
            print("✅ Database initialized successfully")
            sys.stdout.flush()
    except Exception as e:
        print(f"❌ Database init error: {str(e)}")
        sys.stdout.flush()

init_database()

# ─── Excel Price Loader ───────────────────────────────────────────────────────
import os as _os

EXCEL_PATH = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "pricelist.xlsx")

def load_prices_from_excel():
    """
    Read pricelist.xlsx and return a dict:
      {
        'solar_panels': [{'item': ..., 'price': ..., 'installer_price': ...}, ...],
        'pv_mountings': [{'item': ..., 'brand': ..., 'price': ...}, ...],
        'dc_breakers':  [...],
        'ac_breakers':  [...],
        'spd':          [...],
        'mc4':          [...],
        'battery_breaker': [...],
      }
    Returns empty dict on any error.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        data = {}

        # --- Solar Panels ---
        ws = wb["SOLAR PANEL"]
        panels = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            item, brand, _, price, installer = (row + (None,)*5)[:5]
            if item and price:
                item_clean = str(item).replace("\n", " ").strip()
                price_clean = str(price).replace("P", "₱").strip()
                inst_clean  = str(installer).replace("P", "₱").strip() if installer else None
                panels.append({"item": item_clean, "brand": brand, "price": price_clean, "installer_price": inst_clean})
        data["solar_panels"] = panels

        # --- PV Mountings ---
        ws = wb["PV MOUNTINGS"]
        mountings = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            item, brand, _, price = (row + (None,)*4)[:4]
            if item and price is not None:
                mountings.append({"item": str(item).strip(), "brand": brand, "price": f"₱{int(price):,}"})
        data["pv_mountings"] = mountings

        # --- DC Breakers ---
        ws = wb["DC BREAKERS"]
        dc = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            item, brand, _, price = (row + (None,)*4)[:4]
            if item and price is not None:
                dc.append({"item": str(item).strip(), "brand": brand, "price": f"₱{int(price):,}"})
        data["dc_breakers"] = dc

        # --- AC Breakers ---
        ws = wb["AC BREAKERS"]
        ac = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            item, brand, _, price = (row + (None,)*4)[:4]
            if item and price is not None:
                ac.append({"item": str(item).strip(), "brand": brand, "price": f"₱{int(price):,}"})
        data["ac_breakers"] = ac

        # --- SPD ---
        ws = wb["AC & DC SPD"]
        spd = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            item, brand, _, price = (row + (None,)*4)[:4]
            if item and price is not None:
                spd.append({"item": str(item).strip(), "brand": brand, "price": f"₱{int(price):,}"})
        data["spd"] = spd

        # --- MC4 ---
        ws = wb["MC4"]
        mc4 = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            item, brand, _, price = (row + (None,)*4)[:4]
            if item and price is not None:
                mc4.append({"item": str(item).strip(), "brand": brand, "price": f"₱{int(price):,}"})
        data["mc4"] = mc4

        # --- Battery Breaker ---
        ws = wb["BATTERY BREAKER"]
        bb = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            item, brand, _, price = (row + (None,)*4)[:4]
            if item and price is not None:
                bb.append({"item": str(item).strip(), "brand": brand, "price": f"₱{int(price):,}"})
        data["battery_breaker"] = bb

        # --- Conduit ---
        if "CONDUIT" in wb.sheetnames:
            ws = wb["CONDUIT"]
            conduit = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                item, brand, _, price = (row + (None,)*4)[:4]
                if item and price is not None:
                    if isinstance(price, str):
                        import re as _re
                        price_str = _re.sub(r'^P(?=\d)', '₱', str(price).strip())
                    else:
                        price_str = f"₱{int(price):,}/meter"
                    conduit.append({"item": str(item).strip(), "brand": brand or "", "price": price_str})
            data["conduit"] = conduit

        print(f"✅ Price list loaded from Excel: {sum(len(v) for v in data.values())} items")
        sys.stdout.flush()
        return data
    except Exception as e:
        print(f"⚠️  Could not load Excel price list: {e}")
        sys.stdout.flush()
        return {}


def build_pricelist_answer(lang="en"):
    """Build a formatted price list message from the Excel data."""
    prices = load_prices_from_excel()
    if not prices:
        # Fallback to hardcoded if Excel unavailable
        return None

    lines = []
    if lang == "tl":
        lines.append("💰 *Opisyal na Listahan ng Presyo ng Ultiphoton:*")
    else:
        lines.append("💰 *Ultiphoton Official Price List:*")

    # Solar Panels
    if prices.get("solar_panels"):
        lines.append("\n☀️ *Solar Panels (Talesun):*")
        for p in prices["solar_panels"]:
            line = f"- {p['item']}: {p['price']}"
            if p.get("installer_price"):
                line += f" / {p['installer_price']} (installer)"
            lines.append(line)

    # PV Mountings
    if prices.get("pv_mountings"):
        lines.append("\n📌 *PV Mountings (SoEasy):*")
        for p in prices["pv_mountings"]:
            lines.append(f"- {p['item']}: {p['price']}")

    # DC Breakers
    if prices.get("dc_breakers"):
        lines.append("\n📌 *DC Breakers (Chint/Chyt):*")
        for p in prices["dc_breakers"]:
            lines.append(f"- {p['item']}: {p['price']}")

    # AC Breakers
    if prices.get("ac_breakers"):
        lines.append("\n📌 *AC Breakers (Chint/Chyt):*")
        for p in prices["ac_breakers"]:
            lines.append(f"- {p['item']}: {p['price']}")

    # SPD
    if prices.get("spd"):
        lines.append("\n📌 *Surge Protection (SPD):*")
        for p in prices["spd"]:
            lines.append(f"- {p['item']}: {p['price']}")

    # MC4
    if prices.get("mc4"):
        lines.append("\n📌 *Connectors:*")
        for p in prices["mc4"]:
            lines.append(f"- {p['item']}: {p['price']}")

    # Battery Breaker
    if prices.get("battery_breaker"):
        lines.append("\n📌 *Battery Breaker:*")
        for p in prices["battery_breaker"]:
            lines.append(f"- {p['item']}: {p['price']}")

    # Conduit (from Excel or hardcoded fallback)
    conduit_items = prices.get("conduit", [])
    if conduit_items:
        lines.append("\n📌 *Conduit:*")
        for c in conduit_items:
            lines.append(f"- {c['item']}: {c['price']}")
    else:
        if lang == "tl":
            lines.append("\n📌 *Conduit:*")
            lines.append("- HDPE 25mm: ₱170/metro")
        else:
            lines.append("\n📌 *Conduit:*")
            lines.append("- HDPE 25mm: ₱170/meter")

    # PV Cable (hardcoded — not in Excel)
    if lang == "tl":
        lines.append("\n📌 *PV Cable / Wire (Single Core):*")
        lines.append("- 4mm: ₱6,100/roll | ₱70/metro")
        lines.append("- 6mm: ₱7,500/roll | ₱85/metro")
    else:
        lines.append("\n📌 *PV Cable / Wire (Single Core):*")
        lines.append("- 4mm: ₱6,100/roll | ₱70/meter")
        lines.append("- 6mm: ₱7,500/roll | ₱85/meter")

    # Inverters (hardcoded — not in Excel)
    if lang == "tl":
        lines.append("\n⚡ *GoodWe Inverter:*")
        lines.append("- 8 kW: ₱72,000")
        lines.append("- 10 kW: ₱83,000")
        lines.append("- 12 kW: ₱92,000")
    else:
        lines.append("\n⚡ *GoodWe Inverter:*")
        lines.append("- 8 kW: ₱72,000")
        lines.append("- 10 kW: ₱83,000")
        lines.append("- 12 kW: ₱92,000")

    # Batteries (hardcoded — not in Excel)
    if lang == "tl":
        lines.append("\n🔋 *GoodWe Battery:*")
        lines.append("- 280Ah: ₱118,000")
    else:
        lines.append("\n🔋 *GoodWe Battery:*")
        lines.append("- 280Ah: ₱118,000")

    if lang == "tl":
        lines.append("\nMakipag-ugnayan para sa bulk orders at espesyal na presyo! 📞")
    else:
        lines.append("\nContact us for bulk orders & special pricing! 📞")

    return "\n".join(lines)


def build_solar_panel_answer(lang="en"):
    """Build solar panel pricing answer from Excel."""
    prices = load_prices_from_excel()
    panels = prices.get("solar_panels", [])
    if not panels:
        return None

    lines = []
    if lang == "tl":
        lines.append("☀️ *Presyo ng Solar Panels (Talesun):*")
    else:
        lines.append("☀️ *Solar Panel Pricing (Talesun):*")

    for p in panels:
        lines.append(f"\n*{p['item']}*")
        lines.append(f"- Retail: {p['price']}")
        if p.get("installer_price"):
            lines.append(f"- Installer: {p['installer_price']}")

    if lang == "tl":
        lines.append("\nMakipag-ugnayan sa amin para sa bulk orders! 📞")
    else:
        lines.append("\nContact us for bulk orders and special pricing! 📞")

    return "\n".join(lines)


def build_accessories_answer(lang="en"):
    """Build accessories/materials pricing answer from Excel."""
    prices = load_prices_from_excel()
    sections = [
        ("pv_mountings",    "📌 *PV Mountings (SoEasy):*"),
        ("dc_breakers",     "📌 *DC Breakers (Chint/Chyt):*"),
        ("ac_breakers",     "📌 *AC Breakers (Chint/Chyt):*"),
        ("spd",             "📌 *Surge Protection (SPD):*"),
        ("mc4",             "📌 *Connectors:*"),
        ("battery_breaker", "📌 *Battery Breaker:*"),
    ]
    if not any(prices.get(k) for k, _ in sections):
        return None

    lines = []
    if lang == "tl":
        lines.append("🔧 *Listahan ng Presyo ng Accessories & Materials:*")
    else:
        lines.append("🔧 *Accessories & Materials Price List:*")

    for key, header in sections:
        items = prices.get(key, [])
        if items:
            lines.append(f"\n{header}")
            for p in items:
                lines.append(f"- {p['item']}: {p['price']}")

    # Conduit — from Excel if available, else hardcoded fallback
    conduit_items = prices.get("conduit", [])
    if conduit_items:
        lines.append("\n📌 *Conduit:*")
        for c in conduit_items:
            lines.append(f"- {c['item']}: {c['price']}")
    else:
        if lang == "tl":
            lines.append("\n📌 *Conduit:*")
            lines.append("- HDPE 25mm: ₱170/metro")
        else:
            lines.append("\n📌 *Conduit:*")
            lines.append("- HDPE 25mm: ₱170/meter")

    if lang == "tl":
        lines.append("\nMakipag-ugnayan para sa bulk orders! ⚡")
    else:
        lines.append("\nContact us for bulk orders! ⚡")

    return "\n".join(lines)

# ─────────────────────────────────────────────────────────────────────────────

# FAQ Database with Updated Information
FAQS = {
    "solar_panel_price": {
        "keywords": ["magkano", "price", "cost", "solar panel", "talesun", "585w", "620w", "how much", "presyo", "panel price", "solar price", "solar panels"],
        "answer_en": """☀️ **Solar Panel Pricing (Talesun):**

**620W Bifacial**
- Retail: ₱6,100/pc
- Installer price: ₱5,850/pc

**585W Bifacial**
- Retail: ₱5,750/pc
- Installer price: ₱5,650/pc

Contact us for bulk orders and special pricing! 📞""",
        "answer_tl": """☀️ **Presyo ng Solar Panels (Talesun):**

**620W Bifacial**
- Retail: ₱6,100/pc
- Installer price: ₱5,850/pc

**585W Bifacial**
- Retail: ₱5,750/pc
- Installer price: ₱5,650/pc

Makipag-ugnayan sa amin para sa bulk orders! 📞"""
    },
    
    "location": {
        "keywords": ["location", "located", "saan", "address", "office", "branch", "where", "lokasyon"],
        "answer_en": """📍 Our Locations:

Main Office - UltiPhoton Solar Muntinlupa
Filinvest, Muntinlupa City
https://share.google/Qke5XC4NOla3Kq8ef

Branch - UltiPhoton Solar Batangas
Batangas City
https://share.google/0IVMOYIrK7UCIlKFw

Feel free to visit us! ☀️""",
        "answer_tl": """📍 Aming Mga Lokasyon:

Main Office - UltiPhoton Solar Muntinlupa
Filinvest, Muntinlupa City
https://share.google/Qke5XC4NOla3Kq8ef

Branch - UltiPhoton Solar Batangas
Batangas City
https://share.google/0IVMOYIrK7UCIlKFw

Bisitahin kami anumang oras! ☀️"""
    },
    
    "cod": {
        "keywords": ["cod", "delivery", "cash on delivery", "available", "area", "deliver", "delivery", "lalamove", "magpadala", "ipadala", "ship", "shipping", "paano i-deliver", "paano mag-deliver", "how to deliver", "how to ship"],
        "answer_en": """🚚 **Delivery Information:**

For delivery, we recommend using **Lalamove** for fast and reliable service!

Please provide your details so we can arrange the delivery:

📋 **Name:**
📍 **Address:**
📞 **Contact Number:**
🏠 **Nearest Landmark:**

Once we receive your details, we will coordinate the Lalamove booking for you! ☀️""",
        "answer_tl": """🚚 **Impormasyon sa Delivery:**

Para sa delivery, inirerekomenda namin ang **Lalamove** para sa mabilis at maaasahang serbisyo!

Mangyaring ibigay ang inyong mga detalye para maayos namin ang delivery:

📋 **Pangalan:**
📍 **Address:**
📞 **Contact Number:**
🏠 **Pinakamalapit na Landmark:**

Kapag natanggap na namin ang inyong mga detalye, aayusin na namin ang Lalamove booking para sa inyo! ☀️"""
    },
    
    "payment": {
        "keywords": ["payment", "pay", "bank", "transfer", "cash", "gcash", "paano", "bayad", "magbayad"],
        "answer_en": """💳 **Payment Methods:**

We accept **Bank Transfer ONLY** (No cash payments)

**Bank Details:**

**BDO:**
Account Name: JERWIN JEFFREY GAPUZ
Savings Account: 0081 9600 1660

**UnionBank:**
Account Name: JERWIN JEFFREY GAPUZ
Savings Account: 1094 2921 2487

**BPI:**
Account Name: JERWIN JEFFREY GAPUZ
Savings Account: 065 6925 517

**EastWest:**
Account: 200064679316

**GCash:**
0997-369-7123 (J. GAPUZ)

Please transfer and provide proof of payment! 🏦""",
        "answer_tl": """💳 **Paraan ng Pagbabayad:**

Tumatanggap lang kami ng **Bank Transfer** (Walang cash)

**Bank Details:**

**BDO:**
Account Name: JERWIN JEFFREY GAPUZ
Savings Account: 0081 9600 1660

**UnionBank:**
Account Name: JERWIN JEFFREY GAPUZ
Savings Account: 1094 2921 2487

**BPI:**
Account Name: JERWIN JEFFREY GAPUZ
Savings Account: 065 6925 517

**EastWest:**
Account: 200064679316

**GCash:**
0997-369-7123 (J. GAPUZ)

Magpadala ng proof of payment! 🏦"""
    },
    
    "accessories": {
        "keywords": [
            # Only GENERIC accessory keywords here — specific items have their own FAQ entries
            "accessories", "mounting", "mountings", "pv mounting", "pv mountings",
            "soeasy", "chint", "chyt", "materials", "wire", "protection", "device",
            "meron", "mga accessories", "mga materials"
        ],
        "answer_en": """🔧 **Accessories & Materials Price List:**

📌 **PV Mountings (SoEasy Brand):**
- Aluminum Railing 2.4m: ₱600/pc
- L Foot: ₱95/pc
- Mid Clamp: ₱85/pc
- End Clamp: ₱85/pc
- Rail Splicer: ₱85/pc
- PV Grounding Lug: ₱70/pc

📌 **DC Breakers (Chint/Chyt):**
- DC Breaker 20A, 2P: ₱680/pc

📌 **AC Breakers (Chint/Chyt):**
- AC Breaker 40A, 2P: ₱750/pc
- AC Breaker 63A, 2P: ₱750/pc
- AC Breaker 100A, 2P: ₱1,300/pc

📌 **Surge Protection (SPD):**
- DC SPD 1200VDC 40kA: ₱780/pc
- AC SPD 2P 400V: ₱580/pc
- AC SPD 4P 385V: ₱980/pc

📌 **Connectors & Others:**
- MC4 30A DC 1000V (Male & Female pair): ₱80/pair
- Battery Breaker DC 250AT: ₱1,700/pc

📌 **Conduit:**
- HDPE 25mm: ₱175/meter

Contact us for bulk orders! ⚡""",
        "answer_tl": """🔧 **Listahan ng Presyo ng Accessories & Materials:**

📌 **PV Mountings (SoEasy Brand):**
- Aluminum Railing 2.4m: ₱600/pc
- L Foot: ₱95/pc
- Mid Clamp: ₱85/pc
- End Clamp: ₱85/pc
- Rail Splicer: ₱85/pc
- PV Grounding Lug: ₱70/pc

📌 **DC Breakers (Chint/Chyt):**
- DC Breaker 20A, 2P: ₱680/pc

📌 **AC Breakers (Chint/Chyt):**
- AC Breaker 40A, 2P: ₱750/pc
- AC Breaker 63A, 2P: ₱750/pc
- AC Breaker 100A, 2P: ₱1,300/pc

📌 **Surge Protection (SPD):**
- DC SPD 1200VDC 40kA: ₱780/pc
- AC SPD 2P 400V: ₱580/pc
- AC SPD 4P 385V: ₱980/pc

📌 **Connectors & Others:**
- MC4 30A DC 1000V (Male & Female pair): ₱80/pair
- Battery Breaker DC 250AT: ₱1,700/pc

📌 **Conduit:**
- HDPE 25mm: ₱175/metro

Makipag-ugnayan para sa bulk orders! ⚡"""
    },
    
    "inverter_brands": {
        "keywords": ["inverter", "brand", "deye", "solis", "goodwe", "srne", "sigenergy", "ano mga"],
        "answer_en": """⚡ **Inverter Brands Available:**

✅ **Deye** - 5 years warranty
✅ **Solis** - 5 years warranty
✅ **GoodWe** - 5 years warranty
✅ **SRNE** - 5 years warranty
✅ **Sigenergy** - 10 years warranty

📌 **GoodWe Inverter Pricing:**
- 8 kW: ₱72,000
- 10 kW: ₱83,000
- 12 kW: ₱92,000

🔋 **GoodWe Battery:**
- 280Ah: ₱118,000

All brands are high-quality and reliable for Philippine climate! 🌞""",
        "answer_tl": """⚡ **Available Inverter Brands:**

✅ **Deye** - 5 taong warranty
✅ **Solis** - 5 taong warranty
✅ **GoodWe** - 5 taong warranty
✅ **SRNE** - 5 taong warranty
✅ **Sigenergy** - 10 taong warranty

📌 **Presyo ng GoodWe Inverter:**
- 8 kW: ₱72,000
- 10 kW: ₱83,000
- 12 kW: ₱92,000

🔋 **GoodWe Battery:**
- 280Ah: ₱118,000

Lahat ay high-quality at reliable! 🌞"""
    },
    
    "warranty": {
        "keywords": ["warranty", "years", "guarantee", "coverage", "ilang"],
        "answer_en": """✅ **Warranty Coverage:**

**Solar Panels (Talesun):**
10 years warranty

**Inverters:**
- Deye, Solis, SRNE, GoodWe: 5 years
- Sigenergy: 10 years

Quality guaranteed! ☀️""",
        "answer_tl": """✅ **Warranty Coverage:**

**Solar Panels (Talesun):**
10 taong warranty

**Inverters:**
- Deye, Solis, SRNE, GoodWe: 5 taon
- Sigenergy: 10 taon

Garantisadong kalidad! ☀️"""
    },
    
    "battery": {
        "keywords": ["battery", "storage", "backup", "energy storage", "may battery"],
        "answer_en": """🔋 **Battery Storage:**

Yes, we offer batteries **by order only**

We don't keep batteries in stock, but we can source them for you based on your requirements.

Contact us to discuss your battery needs! 📞""",
        "answer_tl": """🔋 **Battery Storage:**

Oo, nag-aalok kami ng batteries **by order lang**

Walang stock kami pero maaari naming kumuha para sa inyo.

Makipag-ugnayan para sa battery needs! 📞"""
    },
    
    "installation_time": {
        "keywords": ["installation", "how long", "days", "time", "duration", "gaano katagal"],
        "answer_en": """⏱️ **Installation Timeline:**

**Duration:** 1 day to 20 days

**Depends on:**
- Site conditions
- System complexity
- Weather conditions

We'll provide a specific timeline after site inspection! 🏗️""",
        "answer_tl": """⏱️ **Installation Timeline:**

**Duration:** 1 hanggang 20 araw

**Depende sa:**
- Site conditions
- System complexity
- Weather conditions

Magbibigay kami ng specific timeline pagkatapos ng inspection! 🏗️"""
    },
    
    "site_inspection": {
        "keywords": ["site inspection", "libre", "free", "survey", "inspect"],
        "answer_en": """🔍 **Site Inspection:**

✅ **FREE Site Inspection!**

We provide complimentary site inspection to assess your property and design the perfect solar system for you.

**Condition:** We hope you'll choose us for installation! 😊

Schedule your free inspection today! 📞""",
        "answer_tl": """🔍 **Site Inspection:**

✅ **LIBRE ang Site Inspection!**

Nag-aalok kami ng free inspection para malaman ang perfect solar system para sa inyo.

**Condition:** Sana kami ang pipiliin ninyo! 😊

Schedule ngayon! 📞"""
    },
    "pallet_quantity": {
        "keywords": [
            "paleta", "pallet", "ilang piraso", "piraso sa isang paleta",
            "ilang panels sa paleta", "panels per pallet", "how many per pallet",
            "ilang pcs", "ilang pieces", "pieces per pallet", "qty per pallet",
            "quantity per pallet", "ilang panel sa paleta", "pcs per pallet"
        ],
        "answer_en": """📦 **Panels Per Pallet:**

- **Talesun 585W** — 37 pcs per pallet
- **Talesun 620W** — 36 pcs per pallet

For bulk orders, contact us for special pricing! ☀️""",
        "answer_tl": """📦 **Ilang Piraso sa Isang Paleta:**

- **Talesun 585W** — 37 piraso bawat paleta
- **Talesun 620W** — 36 piraso bawat paleta

Para sa bulk orders, makipag-ugnayan para sa espesyal na presyo! ☀️"""
    },

    "full_pricelist": {
        "keywords": [
            "pricelist", "price list", "listahan ng presyo", "presyo ng lahat",
            "lahat ng presyo", "complete price", "full price", "all prices",
            "materials price", "magkano lahat", "ano lahat", "list of prices",
            "price ng materials", "presyo ng materials",
            # Solar materials queries
            "solar materials", "mga solar materials", "ano po mga solar materials",
            "ano mga solar materials", "available materials", "mga available",
            "solar products", "mga solar products", "ano ang solar materials",
            "lahat ng solar", "lahat ng materials", "complete list",
            "complete materials", "full materials", "all materials",
            "mga gamit", "ano mga gamit", "lahat ng gamit", "mga available na gamit"
        ],
        "answer_en": """💰 **Ultiphoton Official Price List:**

☀️ **Solar Panels (Talesun):**
- 620W: ₱6,100 retail / ₱5,850 installer
- 585W: ₱5,750 retail / ₱5,650 installer

📌 **PV Mountings (SoEasy):**
- Aluminum Railing 2.4m: ₱600
- L Foot: ₱95 | Mid/End Clamp: ₱85
- Rail Splicer: ₱85 | Grounding Lug: ₱70

📌 **Breakers & Protection (Chint/Chyt):**
- DC Breaker 20A 2P: ₱680
- AC Breaker 40A/63A 2P: ₱750
- AC Breaker 100A 2P: ₱1,300
- DC SPD 1200V 40kA: ₱780
- AC SPD 2P 400V: ₱580
- AC SPD 4P 385V: ₱980

📌 **Connectors:**
- MC4 30A 1000V pair: ₱80
- Battery Breaker DC 250AT: ₱1,700

Contact us for bulk orders & inverter pricing! 📞""",
        "answer_tl": """💰 **Opisyal na Listahan ng Presyo ng Ultiphoton:**

☀️ **Solar Panels (Talesun):**
- 620W: ₱6,100 retail / ₱5,850 installer
- 585W: ₱5,750 retail / ₱5,650 installer

📌 **PV Mountings (SoEasy):**
- Aluminum Railing 2.4m: ₱600
- L Foot: ₱95 | Mid/End Clamp: ₱85
- Rail Splicer: ₱85 | Grounding Lug: ₱70

📌 **Breakers & Protection (Chint/Chyt):**
- DC Breaker 20A 2P: ₱680
- AC Breaker 40A/63A 2P: ₱750
- AC Breaker 100A 2P: ₱1,300
- DC SPD 1200V 40kA: ₱780
- AC SPD 2P 400V: ₱580
- AC SPD 4P 385V: ₱980

📌 **Connectors:**
- MC4 30A 1000V pair: ₱80
- Battery Breaker DC 250AT: ₱1,700

Makipag-ugnayan para sa bulk orders at inverter pricing! 📞"""
    },
    "products_and_price": {
        "keywords": [
            # English phrasings
            "what are your products", "what products", "products and price",
            "products and how much", "what do you sell", "what do you offer",
            "available products", "list of products", "product list",
            "what can i buy", "what items", "what materials",
            "show me your products", "show products", "your products",
            "anong products", "anong meron", "anong available",
            # Tagalog phrasings
            "ano ang mga produkto", "ano ang products", "ano ang ibinebenta",
            "ano ang meron kayo", "anong meron kayo", "ano ang available",
            "mga produkto at presyo", "produkto at presyo", "produkto at magkano",
            "ano ang binebenta", "ano ang ibinebenta ninyo", "anong binebenta",
            "ipakita ang products", "ipakita ang presyo", "lahat ng produkto",
            "mga available na produkto", "anong solar", "anong panels"
        ],
        # answer is built dynamically from Excel via get_faq_answer
        "answer_en": "[excel:full_pricelist]",
        "answer_tl": "[excel:full_pricelist]"
    },
    # --- Per-item accessory FAQs (specific items return only their own price) ---
    "railing": {
        "keywords": ["railing", "railings", "aluminum railing", "rail", "rails", "magkano railing", "presyo ng railing", "railing price", "solar rail", "solar rails", "solar railing", "pv rail", "mounting rail", "panel rail"],
        "answer_en": "Aluminum Railing 2.4m (SoEasy): ₱600/pc ☀️",
        "answer_tl": "Aluminum Railing 2.4m (SoEasy): ₱600/pc ☀️"
    },
    "l_foot": {
        "keywords": ["l foot", "l-foot", "lfoot", "magkano l foot", "presyo ng l foot"],
        "answer_en": "L Foot (SoEasy): ₱95/pc ☀️",
        "answer_tl": "L Foot (SoEasy): ₱95/pc ☀️"
    },
    "mid_clamp": {
        "keywords": ["mid clamp", "mid-clamp", "midclamp", "magkano mid clamp", "presyo ng mid clamp"],
        "answer_en": "Mid Clamp (SoEasy): ₱85/pc ☀️",
        "answer_tl": "Mid Clamp (SoEasy): ₱85/pc ☀️"
    },
    "end_clamp": {
        "keywords": ["end clamp", "end-clamp", "endclamp", "magkano end clamp", "presyo ng end clamp"],
        "answer_en": "End Clamp (SoEasy): ₱85/pc ☀️",
        "answer_tl": "End Clamp (SoEasy): ₱85/pc ☀️"
    },
    "rail_splicer": {
        "keywords": ["rail splicer", "rail-splicer", "railsplicer", "splicer", "magkano splicer", "splice kit", "splicing kit", "rail jointer", "rail joiner", "jointer", "joiner", "rail coupling", "coupling", "splice"],
        "answer_en": "Rail Splicer (SoEasy): ₱85/pc ☀️",
        "answer_tl": "Rail Splicer (SoEasy): ₱85/pc ☀️"
    },
    "grounding_lug": {
        "keywords": ["grounding lug", "grounding-lug", "groundinglug", "grounding", "lug", "pv grounding"],
        "answer_en": "PV Grounding Lug (SoEasy): ₱70/pc ☀️",
        "answer_tl": "PV Grounding Lug (SoEasy): ₱70/pc ☀️"
    },
    "dc_breaker": {
        "keywords": ["dc breaker", "dc-breaker", "dcbreaker", "20a breaker", "20a 2p", "magkano dc breaker", "presyo ng dc breaker"],
        "answer_en": "DC Breaker 20A, 2P (Chint/Chyt): ₱680/pc ☀️",
        "answer_tl": "DC Breaker 20A, 2P (Chint/Chyt): ₱680/pc ☀️"
    },
    "ac_breaker": {
        "keywords": ["ac breaker", "ac-breaker", "acbreaker", "magkano ac breaker", "presyo ng ac breaker", "40a", "63a", "100a"],
        "answer_en": "AC Breakers (Chint/Chyt):\n- 40A, 2P: ₱750/pc\n- 63A, 2P: ₱750/pc\n- 100A, 2P: ₱1,300/pc ☀️",
        "answer_tl": "AC Breakers (Chint/Chyt):\n- 40A, 2P: ₱750/pc\n- 63A, 2P: ₱750/pc\n- 100A, 2P: ₱1,300/pc ☀️"
    },
    "dc_spd": {
        "keywords": ["dc spd", "dc-spd", "dcspd", "1200v spd", "1200vdc", "magkano dc spd", "presyo ng dc spd"],
        "answer_en": "DC SPD 1200VDC 40kA (Chint/Chyt): ₱780/pc ☀️",
        "answer_tl": "DC SPD 1200VDC 40kA (Chint/Chyt): ₱780/pc ☀️"
    },
    "ac_spd": {
        "keywords": ["ac spd", "ac-spd", "acspd", "400v spd", "385v spd", "magkano ac spd", "presyo ng ac spd"],
        "answer_en": "AC SPD (Chint/Chyt):\n- 2P 400V: ₱580/pc\n- 4P 385V: ₱980/pc ☀️",
        "answer_tl": "AC SPD (Chint/Chyt):\n- 2P 400V: ₱580/pc\n- 4P 385V: ₱980/pc ☀️"
    },
    "mc4": {
        "keywords": ["mc4", "mc-4", "connector", "connectors", "male female pair", "magkano mc4", "presyo ng mc4"],
        "answer_en": "MC4 30A DC 1000V Male & Female Pair (Chint/Chyt): ₱80/pair ☀️",
        "answer_tl": "MC4 30A DC 1000V Male & Female Pair (Chint/Chyt): ₱80/pair ☀️"
    },
    "battery_breaker": {
        "keywords": ["battery breaker", "battery-breaker", "250at", "dc 250at", "magkano battery breaker", "presyo ng battery breaker"],
        "answer_en": "Battery Breaker DC 250AT (Chint/Chyt): ₱1,700/pc ☀️",
        "answer_tl": "Battery Breaker DC 250AT (Chint/Chyt): ₱1,700/pc ☀️"
    },
    "pv_cable": {
        "keywords": [
            "pv cable", "pv wire", "cable", "wire", "solar wire", "solar cable",
            "4mm", "6mm", "4mm wire", "6mm wire", "4mm cable", "6mm cable",
            "meron cable", "meron wire", "may cable", "may wire",
            "magkano cable", "magkano wire", "presyo ng cable", "presyo ng wire",
            "pv cable price", "wire price"
        ],
        "answer_en": """✅ Yes! Single Core PV cables are available!

📌 **PV Cable / Wire — Single Core (per roll):**
- 4mm: ₱6,100/roll
- 6mm: ₱7,500/roll

📌 **PV Cable / Wire — Single Core (per meter):**
- 4mm: ₱70/meter
- 6mm: ₱85/meter

Contact us to order! ☀️""",
        "answer_tl": """✅ Yes! Available po ang Single Core PV cables!

📌 **PV Cable / Wire — Single Core (bawat roll):**
- 4mm: ₱6,100/roll
- 6mm: ₱7,500/roll

📌 **PV Cable / Wire — Single Core (bawat metro):**
- 4mm: ₱70/metro
- 6mm: ₱85/metro

Makipag-ugnayan para mag-order! ☀️"""
    },

    "hdpe_price": {
        "keywords": [
            "hdpe", "hdpe 25mm", "hdpe conduit", "hdpe pipe", "hdpe tubing",
            "25mm hdpe", "hdpe 25", "conduit", "magkano hdpe", "presyo ng hdpe",
            "hdpe price", "hdpe cost", "how much hdpe", "meron hdpe"
        ],
        "answer_en": """📌 **HDPE Conduit 25mm:**
- ₱170/meter

Available po! Contact us to order! ☀️""",
        "answer_tl": """📌 **HDPE Conduit 25mm:**
- ₱170/metro

Available po! Makipag-ugnayan para mag-order! ☀️"""
    },

    # ── Installation Estimate FAQs (per kW, On-Grid & Hybrid) ──────────────────
    "install_3kw": {
        "keywords": [
            "3kw", "3 kw", "3kilowatt", "3 kilowatt", "3kva", "3 kva",
            "3kw system", "3kw solar", "3kw installation", "3kw setup",
            "magkano 3kw", "presyo ng 3kw", "estimate 3kw", "quote 3kw",
            "how much 3kw", "3kw price", "3kw cost",
            "3kw setup", "magkano 3kw setup", "how much 3kw setup",
            "3 kw setup", "3kw package", "3kw quote"
        ],
        "answer_en": """☀️ **3 kW Solar System — Supply & Installation Estimate:**

⚡ **On-Grid (Grid-Tied, No Battery):**
- Estimated Cost: ₱160,000 – ₱235,000
- Best for: Lower upfront cost, faster ROI (4–7 years)
- Includes: Solar panels, grid-tie inverter, mounting, wiring, labor & basic permits

🔋 **Hybrid (With Battery Backup):**
- Estimated Cost: ₱290,000 – ₱380,000
- Best for: Backup during brownouts & night use
- Includes: Solar panels, hybrid inverter, lithium battery, mounting, wiring & full installation

📌 *Prices may vary depending on brand, roof type, and installation complexity.*
Contact us for a free site assessment! 💚""",
        "answer_tl": """☀️ **3 kW Solar System — Estimate ng Presyo (Supply & Installation):**

⚡ **On-Grid (Walang Battery):**
- Estimated na Halaga: ₱160,000 – ₱235,000
- Para sa: Mas mababang gastos, mas mabilis na ROI (4–7 taon)
- Kasama: Solar panels, grid-tie inverter, mounting, wiring, labor at basic permits

🔋 **Hybrid (May Battery Backup):**
- Estimated na Halaga: ₱290,000 – ₱380,000
- Para sa: Backup tuwing may brownout at gabi
- Kasama: Solar panels, hybrid inverter, lithium battery, mounting, wiring at buong installation

📌 *Maaaring mag-iba ang presyo depende sa brand, uri ng bubong, at complexity ng installation.*
Makipag-ugnayan para sa libreng site assessment! 💚"""
    },
    "install_5kw": {
        "keywords": [
            "5kw", "5 kw", "5kilowatt", "5 kilowatt", "5kva", "5 kva",
            "5kw system", "5kw solar", "5kw installation", "5kw setup",
            "magkano 5kw", "presyo ng 5kw", "estimate 5kw", "quote 5kw",
            "how much 5kw", "5kw price", "5kw cost",
            "5kw setup", "magkano 5kw setup", "how much 5kw setup",
            "5 kw setup", "5kw package", "5kw quote"
        ],
        "answer_en": """☀️ **5 kW Solar System — Supply & Installation Estimate:**

⚡ **On-Grid (Grid-Tied, No Battery):**
- Estimated Cost: ₱280,000 – ₱380,000
- Best for: Lower upfront cost, faster ROI (4–7 years)
- Includes: Solar panels, grid-tie inverter, mounting, wiring, labor & basic permits

🔋 **Hybrid (With Battery Backup):**
- Estimated Cost: ₱375,000 – ₱480,000
- Best for: Backup during brownouts & night use
- Includes: Solar panels, hybrid inverter, lithium battery, mounting, wiring & full installation

📌 *Prices may vary depending on brand, roof type, and installation complexity.*
Contact us for a free site assessment! 💚""",
        "answer_tl": """☀️ **5 kW Solar System — Estimate ng Presyo (Supply & Installation):**

⚡ **On-Grid (Walang Battery):**
- Estimated na Halaga: ₱280,000 – ₱380,000
- Para sa: Mas mababang gastos, mas mabilis na ROI (4–7 taon)
- Kasama: Solar panels, grid-tie inverter, mounting, wiring, labor at basic permits

🔋 **Hybrid (May Battery Backup):**
- Estimated na Halaga: ₱375,000 – ₱480,000
- Para sa: Backup tuwing may brownout at gabi
- Kasama: Solar panels, hybrid inverter, lithium battery, mounting, wiring at buong installation

📌 *Maaaring mag-iba ang presyo depende sa brand, uri ng bubong, at complexity ng installation.*
Makipag-ugnayan para sa libreng site assessment! 💚"""
    },
    "install_8kw": {
        "keywords": [
            "8kw", "8 kw", "8kilowatt", "8 kilowatt", "8kva", "8 kva",
            "8kw system", "8kw solar", "8kw installation", "8kw setup",
            "magkano 8kw", "presyo ng 8kw", "estimate 8kw", "quote 8kw",
            "how much 8kw", "8kw price", "8kw cost",
            "8kw setup", "magkano 8kw setup", "how much 8kw setup",
            "8 kw setup", "8kw package", "8kw quote"
        ],
        "answer_en": """☀️ **8 kW Solar System — Supply & Installation Estimate:**

⚡ **On-Grid (Grid-Tied, No Battery):**
- Estimated Cost: ₱380,000 – ₱500,000
- Best for: Lower upfront cost, faster ROI (4–7 years)
- Includes: Solar panels, grid-tie inverter, mounting, wiring, labor & basic permits

🔋 **Hybrid (With Battery Backup):**
- Estimated Cost: ₱400,000 – ₱700,000
- Best for: Backup during brownouts & night use
- Includes: Solar panels, hybrid inverter, lithium battery, mounting, wiring & full installation

📌 *Prices may vary depending on brand, roof type, and installation complexity.*
Contact us for a free site assessment! 💚""",
        "answer_tl": """☀️ **8 kW Solar System — Estimate ng Presyo (Supply & Installation):**

⚡ **On-Grid (Walang Battery):**
- Estimated na Halaga: ₱380,000 – ₱500,000
- Para sa: Mas mababang gastos, mas mabilis na ROI (4–7 taon)
- Kasama: Solar panels, grid-tie inverter, mounting, wiring, labor at basic permits

🔋 **Hybrid (May Battery Backup):**
- Estimated na Halaga: ₱400,000 – ₱700,000
- Para sa: Backup tuwing may brownout at gabi
- Kasama: Solar panels, hybrid inverter, lithium battery, mounting, wiring at buong installation

📌 *Maaaring mag-iba ang presyo depende sa brand, uri ng bubong, at complexity ng installation.*
Makipag-ugnayan para sa libreng site assessment! 💚"""
    },
    "install_10kw": {
        "keywords": [
            "10kw", "10 kw", "10kilowatt", "10 kilowatt", "10kva", "10 kva",
            "10kw system", "10kw solar", "10kw installation", "10kw setup",
            "magkano 10kw", "presyo ng 10kw", "estimate 10kw", "quote 10kw",
            "how much 10kw", "10kw price", "10kw cost",
            "10kw setup", "magkano 10kw setup", "how much 10kw setup",
            "10 kw setup", "10kw package", "10kw quote"
        ],
        "answer_en": """☀️ **10 kW Solar System — Supply & Installation Estimate:**

⚡ **On-Grid (Grid-Tied, No Battery):**
- Estimated Cost: ₱420,000 – ₱600,000
- Best for: Lower upfront cost, faster ROI (4–7 years)
- Includes: Solar panels, grid-tie inverter, mounting, wiring, labor & basic permits

🔋 **Hybrid (With Battery Backup):**
- Estimated Cost: ₱650,000 – ₱950,000
- Best for: Backup during brownouts & night use
- Includes: Solar panels, hybrid inverter, lithium battery, mounting, wiring & full installation

📌 *Prices may vary depending on brand, roof type, and installation complexity.*
Contact us for a free site assessment! 💚""",
        "answer_tl": """☀️ **10 kW Solar System — Estimate ng Presyo (Supply & Installation):**

⚡ **On-Grid (Walang Battery):**
- Estimated na Halaga: ₱420,000 – ₱600,000
- Para sa: Mas mababang gastos, mas mabilis na ROI (4–7 taon)
- Kasama: Solar panels, grid-tie inverter, mounting, wiring, labor at basic permits

🔋 **Hybrid (May Battery Backup):**
- Estimated na Halaga: ₱650,000 – ₱950,000
- Para sa: Backup tuwing may brownout at gabi
- Kasama: Solar panels, hybrid inverter, lithium battery, mounting, wiring at buong installation

📌 *Maaaring mag-iba ang presyo depende sa brand, uri ng bubong, at complexity ng installation.*
Makipag-ugnayan para sa libreng site assessment! 💚"""
    },
    "install_12kw": {
        "keywords": [
            "12kw", "12 kw", "12kilowatt", "12 kilowatt", "12kva", "12 kva",
            "12kw system", "12kw solar", "12kw installation", "12kw setup",
            "magkano 12kw", "presyo ng 12kw", "estimate 12kw", "quote 12kw",
            "how much 12kw", "12kw price", "12kw cost",
            "12kw setup", "magkano 12kw setup", "how much 12kw setup",
            "12 kw setup", "12kw package", "12kw quote"
        ],
        "answer_en": """☀️ **12 kW Solar System — Supply & Installation Estimate:**

⚡ **On-Grid (Grid-Tied, No Battery):**
- Estimated Cost: ₱500,000 – ₱750,000
- Best for: Lower upfront cost, faster ROI (4–7 years)
- Includes: Solar panels, grid-tie inverter, mounting, wiring, labor & basic permits

🔋 **Hybrid (With Battery Backup):**
- Estimated Cost: ₱800,000 – ₱1,100,000
- Best for: Backup during brownouts & night use
- Includes: Solar panels, hybrid inverter, lithium battery, mounting, wiring & full installation

📌 *Prices may vary depending on brand, roof type, and installation complexity.*
Contact us for a free site assessment! 💚""",
        "answer_tl": """☀️ **12 kW Solar System — Estimate ng Presyo (Supply & Installation):**

⚡ **On-Grid (Walang Battery):**
- Estimated na Halaga: ₱500,000 – ₱750,000
- Para sa: Mas mababang gastos, mas mabilis na ROI (4–7 taon)
- Kasama: Solar panels, grid-tie inverter, mounting, wiring, labor at basic permits

🔋 **Hybrid (May Battery Backup):**
- Estimated na Halaga: ₱800,000 – ₱1,100,000
- Para sa: Backup tuwing may brownout at gabi
- Kasama: Solar panels, hybrid inverter, lithium battery, mounting, wiring at buong installation

📌 *Maaaring mag-iba ang presyo depende sa brand, uri ng bubong, at complexity ng installation.*
Makipag-ugnayan para sa libreng site assessment! 💚"""
    },
    "install_15kw": {
        "keywords": [
            "15 kw", "15kilowatt", "15 kilowatt", "15kva", "15 kva",
            "16kw", "16 kw", "15-16kw", "15 to 16kw",
            "15kw system", "15kw solar", "15kw installation", "15kw setup",
            "magkano 15kw", "presyo ng 15kw", "estimate 15kw", "quote 15kw",
            "how much 15kw", "15kw price", "15kw cost",
            "magkano 16kw", "16kw price", "16kw cost",
            "15kw setup", "magkano 15kw setup", "how much 15kw setup",
            "16kw setup", "magkano 16kw setup", "15 kw setup", "16 kw setup"
        ],
        "answer_en": """☀️ **15–16 kW Solar System — Supply & Installation Estimate:**

⚡ **On-Grid (Grid-Tied, No Battery):**
- Estimated Cost: ₱600,000 – ₱1,000,000
- Best for: Large homes / heavy AC usage, faster ROI
- Includes: Solar panels, grid-tie inverter, mounting, wiring, labor & basic permits

🔋 **Hybrid (With Battery Backup):**
- Estimated Cost: ₱950,000 – ₱1,500,000+
- Best for: Backup during brownouts & night use
- Includes: Solar panels, hybrid inverter, lithium battery, mounting, wiring & full installation

📌 *Prices may vary depending on brand, roof type, and installation complexity.*
Contact us for a free site assessment! 💚""",
        "answer_tl": """☀️ **15–16 kW Solar System — Estimate ng Presyo (Supply & Installation):**

⚡ **On-Grid (Walang Battery):**
- Estimated na Halaga: ₱600,000 – ₱1,000,000
- Para sa: Malalaking bahay / mabigat na AC usage, mas mabilis na ROI
- Kasama: Solar panels, grid-tie inverter, mounting, wiring, labor at basic permits

🔋 **Hybrid (May Battery Backup):**
- Estimated na Halaga: ₱950,000 – ₱1,500,000+
- Para sa: Backup tuwing may brownout at gabi
- Kasama: Solar panels, hybrid inverter, lithium battery, mounting, wiring at buong installation

📌 *Maaaring mag-iba ang presyo depende sa brand, uri ng bubong, at complexity ng installation.*
Makipag-ugnayan para sa libreng site assessment! 💚"""
    },
    "install_6kw": {
        "keywords": [
            "6kw", "6 kw", "6kilowatt", "6 kilowatt", "6kva", "6 kva",
            "6kw system", "6kw solar", "6kw installation", "6kw setup",
            "magkano 6kw", "presyo ng 6kw", "estimate 6kw", "quote 6kw",
            "how much 6kw", "6kw price", "6kw cost",
            "6kw setup", "magkano 6kw setup", "how much 6kw setup",
            "6 kw setup", "6kw package", "6kw quote"
        ],
        "answer_en": """☀️ **6 kW Solar System — Supply & Installation Estimate:**

⚡ **On-Grid (Grid-Tied, No Battery):**
- Estimated Cost: ₱300,000 – ₱420,000
- Best for: Lower upfront cost, faster ROI (4–7 years)
- Includes: Solar panels, grid-tie inverter, mounting, wiring, labor & basic permits

🔋 **Hybrid (With Battery Backup):**
- Estimated Cost: ₱420,000 – ₱600,000
- Best for: Backup during brownouts & night use
- Includes: Solar panels, hybrid inverter, lithium battery, mounting, wiring & full installation

📌 *Prices may vary depending on brand, roof type, and installation complexity.*
Contact us for a free site assessment! 💚""",
        "answer_tl": """☀️ **6 kW Solar System — Estimate ng Presyo (Supply & Installation):**

⚡ **On-Grid (Walang Battery):**
- Estimated na Halaga: ₱300,000 – ₱420,000
- Para sa: Mas mababang gastos, mas mabilis na ROI (4–7 taon)
- Kasama: Solar panels, grid-tie inverter, mounting, wiring, labor at basic permits

🔋 **Hybrid (May Battery Backup):**
- Estimated na Halaga: ₱420,000 – ₱600,000
- Para sa: Backup tuwing may brownout at gabi
- Kasama: Solar panels, hybrid inverter, lithium battery, mounting, wiring at buong installation

📌 *Maaaring mag-iba ang presyo depende sa brand, uri ng bubong, at complexity ng installation.*
Makipag-ugnayan para sa libreng site assessment! 💚"""
    },
    "install_general": {
        "keywords": [
            # Generic install / setup queries
            "solar setup", "solar set up", "full package", "full solar package",
            "supply and install", "supply and installation", "supply install",
            "installation estimate", "installation cost estimate",
            "magkano mag-install", "magkano ang installation", "presyo ng installation",
            "estimate ng solar", "solar system price", "solar system cost",
            "on-grid price", "hybrid price", "on grid price", "hybrid system price",
            "solar setup price", "solar setup cost",
            "how much solar", "magkano solar system", "magkano ang solar",
            "complete solar", "complete system", "buong sistema",
            "solar package", "solar installation price", "solar installation cost",
            "magkano ang solar setup", "how much solar setup",
            "full solar", "buong solar", "solar system estimate"
        ],
        "answer_en": """☀️ **PH Solar System Price Estimate (Supply & Installation)**

Scope: Supply, Delivery, Installation, Testing & Commissioning

⚡ **1. ON-GRID System (No Battery — Net Metering Ready)**
Typical cost: ₱40,000 – ₱60,000 per kW

- 3 kW:   ₱160,000 – ₱235,000
- 5 kW:   ₱280,000 – ₱380,000
- 6 kW:   ₱300,000 – ₱420,000
- 8 kW:   ₱380,000 – ₱500,000
- 10 kW:  ₱420,000 – ₱600,000
- 12 kW:  ₱500,000 – ₱750,000
- 15–16 kW: ₱600,000 – ₱1,000,000
✔ Includes: Solar panels, grid-tie inverter, mounting, wiring, labor & basic permits

🔋 **2. HYBRID System (With Battery Backup)**
Typical cost: ₱70,000 – ₱100,000 per kW

- 3 kW:   ₱290,000 – ₱380,000
- 5 kW:   ₱375,000 – ₱480,000
- 6 kW:   ₱420,000 – ₱600,000
- 8 kW:   ₱400,000 – ₱700,000
- 10 kW:  ₱650,000 – ₱950,000
- 12 kW:  ₱800,000 – ₱1,100,000
- 15–16 kW: ₱950,000 – ₱1,500,000+
✔ Includes: Solar panels, hybrid inverter, lithium battery (5–15 kWh), mounting, wiring & full installation

📌 *Prices vary by brand, roof type & installation complexity.*
Ask about a specific size (e.g. "6kW setup") for a detailed quote!
Contact us for a FREE site assessment! 💚""",
        "answer_tl": """☀️ **Estimate ng Presyo ng Solar System sa Pilipinas (Supply & Installation)**

Saklaw: Supply, Delivery, Installation, Testing at Commissioning

⚡ **1. ON-GRID System (Walang Battery — Net Metering Ready)**
Tipikal na halaga: ₱40,000 – ₱60,000 bawat kW

- 3 kW:   ₱160,000 – ₱235,000
- 5 kW:   ₱280,000 – ₱380,000
- 6 kW:   ₱300,000 – ₱420,000
- 8 kW:   ₱380,000 – ₱500,000
- 10 kW:  ₱420,000 – ₱600,000
- 12 kW:  ₱500,000 – ₱750,000
- 15–16 kW: ₱600,000 – ₱1,000,000
✔ Kasama: Solar panels, grid-tie inverter, mounting, wiring, labor at basic permits

🔋 **2. HYBRID System (May Battery Backup)**
Tipikal na halaga: ₱70,000 – ₱100,000 bawat kW

- 3 kW:   ₱290,000 – ₱380,000
- 5 kW:   ₱375,000 – ₱480,000
- 6 kW:   ₱420,000 – ₱600,000
- 8 kW:   ₱400,000 – ₱700,000
- 10 kW:  ₱650,000 – ₱950,000
- 12 kW:  ₱800,000 – ₱1,100,000
- 15–16 kW: ₱950,000 – ₱1,500,000+
✔ Kasama: Solar panels, hybrid inverter, lithium battery (5–15 kWh), mounting, wiring at buong installation

📌 *Maaaring mag-iba ang presyo depende sa brand, uri ng bubong at complexity ng installation.*
Magtanong ng specific na size (hal. "6kW setup") para sa detalyadong quote!
Makipag-ugnayan para sa LIBRENG site assessment! 💚"""
    },
    "website": {
        "keywords": [
            "website", "web site", "online", "link", "url", "site nyo", "webpage",
            "website nyo", "website ninyo", "website niyo", "website po", "website kayo",
            "ano website", "ano po website", "may website", "meron website",
            "your website", "your site", "your page", "your link",
            "ultiphoton.com", "ultiphotonsolar"
        ],
        "answer_en": """🌐 **Our Website:**

Visit us online at:
👉 https://ultiphotonsolarpoweropc.com/

You can find more information about our products, services, and solar solutions on our website! ☀️""",
        "answer_tl": """🌐 **Ang Aming Website:**

Bisitahin kami online sa:
👉 https://ultiphotonsolarpoweropc.com/

Makikita ninyo doon ang lahat ng impormasyon tungkol sa aming mga produkto at serbisyo! ☀️"""
    }
}

# ── Price Calculator ─────────────────────────────────────────────────────────
# Unit prices for all items the bot knows about (in PHP)
# ── Lalamove Delivery Fee Estimation ──────────────────────────────────────────
# Origin: UltiPhoton Solar Batangas branch (used as pickup point)
LALAMOVE_ORIGIN = "Batangas City, Batangas, Philippines"

# Lalamove PH published rates (Manila, NCR & South Luzon region)
# Source: https://www.lalamove.com/en-ph/all-delivery-pricing-detail (verified 2025)
LALAMOVE_RATES = [
    # Motorcycle: ₱49 base + ₱6/km (0-5km) + ₱5/km (above 5km)
    {"vehicle": "Motorcycle (up to 20kg)",  "base": 49,  "r1": 6,  "t1": 5,  "r2": 5,  "max_kg": 20},
    # Sedan 200kg: ₱100 base + ₱18/km (0-5km) + ₱15/km (above 5km)
    {"vehicle": "Sedan 200kg",              "base": 100, "r1": 18, "t1": 5,  "r2": 15, "max_kg": 200},
    # 300kg Subcompact SUV: ₱115 base + ₱20/km (1-30km) + ₱17/km (above 30km)
    {"vehicle": "SUV/Crossover 300kg",      "base": 115, "r1": 20, "t1": 30, "r2": 17, "max_kg": 300},
    # 600kg 7-seater SUV/Small Van: ₱200 base + ₱20/km (1-30km) + ₱17/km (above 30km)
    {"vehicle": "MPV/Small Van 600kg",      "base": 200, "r1": 20, "t1": 30, "r2": 17, "max_kg": 600},
    # 800kg Pickup: ₱240 base + ₱20/km
    {"vehicle": "Pickup 800kg",             "base": 240, "r1": 20, "t1": 9999, "r2": 20, "max_kg": 800},
    # L300/Cargo Van 1000kg: ₱280 base + ₱20/km
    {"vehicle": "L300/Cargo Van 1000kg",    "base": 280, "r1": 20, "t1": 9999, "r2": 20, "max_kg": 1000},
]

# Approximate driving distances (km) from Batangas City (Brgy. Bolbok) to common PH destinations
# Used when Google Maps API key is not available
BATANGAS_DISTANCE_TABLE = {
    # Batangas Province
    "batangas city": 5, "bolbok": 2, "batangas": 5,
    "lipa": 22, "lipa city": 22,
    "sto tomas": 35, "santo tomas": 35,
    "tanauan": 40, "tanauan city": 40,
    "malvar": 38, "balete": 28, "agoncillo": 30,
    "nasugbu": 65, "lemery": 45, "san jose": 55,
    "calaca": 48, "balayan": 52, "calatagan": 70,
    "lobo": 60, "mabini": 55, "tingloy": 65,
    "rosario": 35, "san juan": 50, "taysan": 42,
    "ibaan": 30, "san nicolas": 25, "padre garcia": 38,
    "cuenca": 32, "alitagtag": 35, "laurel": 45,
    "mataas na kahoy": 48, "tuy": 58,
    # Quezon Province
    "tayabas": 85, "tayabas city": 85,
    "lucena": 95, "lucena city": 95,
    "sariaya": 90, "candelaria": 100,
    "tiaong": 75, "rosario quezon": 80,
    "pagbilao": 105, "atimonan": 115,
    # Laguna
    "calamba": 55, "calamba city": 55,
    "sta rosa": 65, "santa rosa": 65,
    "binan": 70, "san pedro": 80,
    "cabuyao": 68, "bay": 72,
    "los banos": 78, "los baños": 78,
    "pagsanjan": 90, "paete": 95,
    "san pablo": 60, "san pablo city": 60,
    # Cavite
    "tagaytay": 45, "silang": 50,
    "dasmariñas": 70, "dasmarinas": 70,
    "bacoor": 90, "imus": 88,
    "cavite city": 95, "general trias": 75,
    "trece martires": 65, "naic": 72,
    # Metro Manila
    "manila": 115, "manila city": 115,
    "makati": 118, "makati city": 118,
    "taguig": 112, "pasay": 116,
    "paranaque": 110, "las pinas": 108,
    "muntinlupa": 100, "alabang": 102,
    "pasig": 122, "mandaluyong": 120,
    "quezon city": 130, "caloocan": 135,
    "marikina": 128, "valenzuela": 140,
    "malabon": 138, "navotas": 140,
    "san juan metro": 122, "pateros": 120,
}

def _estimate_km_from_address(address):
    """Estimate driving distance from Batangas City using the lookup table.
    Returns km (float) or None if address not recognized."""
    addr_low = address.lower()
    # Try longest match first
    best_km = None
    best_len = 0
    for place, km in BATANGAS_DISTANCE_TABLE.items():
        if place in addr_low and len(place) > best_len:
            best_km = km
            best_len = len(place)
    return best_km

def _calc_lalamove_fare(rate, km):
    """Calculate fare for a given rate config and distance in km."""
    tier1_km = min(km, rate["t1"])
    tier2_km = max(0, km - rate["t1"])
    return rate["base"] + tier1_km * rate["r1"] + tier2_km * rate["r2"]

def _geocode_address(address):
    """Return (lat, lng) for an address using Google Maps Geocoding API."""
    if not GOOGLE_MAPS_API_KEY:
        return None
    try:
        url = "https://maps.googleapis.com/maps/api/geocode/json"
        resp = requests.get(url, params={"address": address, "key": GOOGLE_MAPS_API_KEY}, timeout=8)
        data = resp.json()
        if data.get("status") == "OK":
            loc = data["results"][0]["geometry"]["location"]
            return loc["lat"], loc["lng"]
    except Exception as e:
        print(f"❌ Geocode error: {e}")
    return None

def _distance_km(origin_addr, dest_addr):
    """Return driving distance in km between two addresses using Google Distance Matrix."""
    if not GOOGLE_MAPS_API_KEY:
        return None
    try:
        url = "https://maps.googleapis.com/maps/api/distancematrix/json"
        resp = requests.get(url, params={
            "origins": origin_addr,
            "destinations": dest_addr,
            "key": GOOGLE_MAPS_API_KEY,
            "mode": "driving",
            "region": "ph"
        }, timeout=10)
        data = resp.json()
        if data.get("status") == "OK":
            element = data["rows"][0]["elements"][0]
            if element.get("status") == "OK":
                return element["distance"]["value"] / 1000.0  # meters → km
    except Exception as e:
        print(f"❌ Distance Matrix error: {e}")
    return None

# Regex to detect customer delivery detail messages
_ADDR_PATTERN = re.compile(
    r'(?:address|address:|address\s*:|add:|add\s*:|lugar|tirahan|lokasyon|location|brgy|barangay|st\.|street|ave\.|avenue|blk|block|lot|subd|subdivision)'
    r'|(?:name:|name\s*:|pangalan:|pangalan\s*:|cp\s*no|cp\s*:|contact\s*no|contact\s*number|cellphone|mobile)',
    re.IGNORECASE
)

def _extract_address_from_message(message):
    """Try to extract a delivery address from a customer message.
    Returns the address string or None."""
    lines = [l.strip() for l in message.replace(',\n', '\n').split('\n') if l.strip()]
    address_line = None
    for line in lines:
        low = line.lower()
        # Look for address-labelled lines
        if re.match(r'^(address|add|tirahan|lugar|lokasyon)\s*[:\-]?\s*', low):
            # Strip the label
            address_line = re.sub(r'^(address|add|tirahan|lugar|lokasyon)\s*[:\-]?\s*', '', line, flags=re.IGNORECASE).strip()
            break
        # Lines that look like a Philippine address (contain brgy/block/lot/city/province)
        if re.search(r'\b(brgy|barangay|blk|block|lot|city|province|quezon|batangas|laguna|manila|muntinlupa|cavite|rizal|bulacan|pampanga|cebu|davao)\b', low):
            address_line = line
            break
    return address_line if address_line and len(address_line) > 8 else None

def estimate_lalamove_from_message(message, language="en"):
    """Detect if a message contains a delivery address and return a Lalamove fee estimate.
    Returns formatted response string or None if not a delivery address message."""
    # Only trigger if message looks like delivery details
    if not _ADDR_PATTERN.search(message):
        return None
    address = _extract_address_from_message(message)
    if not address:
        return None

    # Try to get driving distance via Google Maps API first, then fallback to lookup table
    km = _distance_km(LALAMOVE_ORIGIN, address + ", Philippines")
    km_source = "exact"
    if km is None:
        km = _estimate_km_from_address(address)
        km_source = "estimated"

    if km is None:
        # Address not recognized in lookup table either — acknowledge and guide
        if language == "tl":
            return (
                f"✅ Natanggap namin ang inyong mga detalye!\n\n"
                f"📍 **Delivery Address:** {address}\n\n"
                f"🚚 Para sa tumpak na presyo ng Lalamove, buksan po ang Lalamove app at i-input ang:\n"
                f"• Pickup: UltiPhoton Solar — Brgy. Bolbok, Batangas City\n"
                f"• Drop-off: {address}\n\n"
                f"Makipag-ugnayan sa amin para ma-confirm ang inyong order! 📞"
            )
        else:
            return (
                f"✅ We received your delivery details!\n\n"
                f"📍 **Delivery Address:** {address}\n\n"
                f"🚚 For the exact Lalamove rate, please open the Lalamove app and enter:\n"
                f"• Pickup: UltiPhoton Solar — Brgy. Bolbok, Batangas City\n"
                f"• Drop-off: {address}\n\n"
                f"Contact us to confirm your order! 📞"
            )

    # Calculate estimates for each vehicle type
    estimates = []
    for rate in LALAMOVE_RATES:
        fare = _calc_lalamove_fare(rate, km)
        estimates.append((rate["vehicle"], fare))

    km_display = f"{km:.1f}"
    if language == "tl":
        lines = [
            f"✅ Natanggap namin ang inyong mga detalye!",
            f"",
            f"📍 **Delivery Address:** {address}",
            f"📏 **Estimated Distance:** ~{km_display} km mula Batangas City",
            f"",
            f"🚚 **Estimated Lalamove Delivery Fee:**",
        ]
        for vehicle, fare in estimates:
            lines.append(f"• {vehicle}: ~₱{fare:,.0f}")
        lines += [
            f"",
            f"⚠️ *Ang presyo ay estimate lamang. Maaaring mag-iba depende sa traffic, surcharge, at Lalamove app.*",
            f"",
            f"Makipag-ugnayan sa amin para ma-confirm ang inyong order! 📞"
        ]
    else:
        lines = [
            f"✅ We received your delivery details!",
            f"",
            f"📍 **Delivery Address:** {address}",
            f"📏 **Estimated Distance:** ~{km_display} km from Batangas City",
            f"",
            f"🚚 **Estimated Lalamove Delivery Fee:**",
        ]
        for vehicle, fare in estimates:
            lines.append(f"• {vehicle}: ~₱{fare:,.0f}")
        lines += [
            f"",
            f"⚠️ *Prices are estimates only. Actual fare may vary based on traffic, surcharges, and Lalamove app pricing.*",
            f"",
            f"Contact us to confirm your order! 📞"
        ]
    return "\n".join(lines)

# ── End Lalamove Delivery Fee Estimation ───────────────────────────────────────

UNIT_PRICES = {
    # PV Mountings (SoEasy)
    "railing":        {"price": 600,   "unit": "pc",   "aliases": ["railings", "aluminum railing", "alum railing", "aluminium railing", "2.4m rail", "2.4m railing", "rail 2.4m", "rail", "rails", "solar rail", "solar rails", "solar railing", "solar railings", "pv rail", "pv rails", "mounting rail", "mounting rails", "panel rail", "panel rails"]},
    "l-foot":         {"price": 95,    "unit": "pc",   "aliases": ["l foot", "lfoot", "l-feet", "l feet", "lfeet", "l-foots", "l clamp", "l-clamp", "lclamp", "l clamps", "l-clamps", "lft", "l ft", "l-ft", "lf"]},
    "mid clamp":      {"price": 85,    "unit": "pc",   "aliases": ["midclamp", "mid-clamp", "middle clamp", "mid clamps"]},
    "end clamp":      {"price": 85,    "unit": "pc",   "aliases": ["endclamp", "end-clamp", "end clamps", "end connector", "end connectors"]},
    "rail splicer":   {"price": 85,    "unit": "pc",   "aliases": ["rail splice", "splicer", "rail connector", "splicer connector", "splice connector", "splicing connector", "splice kit", "splicing kit", "rail jointer", "rail joint", "rail joiner", "jointer", "joiner", "rail coupling", "coupling", "rail splice kit", "splice"]},
    "pv grounding lug": {"price": 70, "unit": "pc",   "aliases": ["grounding lug", "grounding lugs", "ground lug", "ground lugs", "pv lug", "pv lugs", "grounding"]},
    # DC Breakers
    "dc breaker":     {"price": 680,   "unit": "pc",   "aliases": ["dc circuit breaker", "dc cb", "dc breakers", "dc braker"]},
    # AC Breakers
    "ac breaker 40a": {"price": 750,   "unit": "pc",   "aliases": ["40a breaker", "40amp breaker", "40a ac breaker", "40 amp breaker"]},
    "ac breaker 63a": {"price": 750,   "unit": "pc",   "aliases": ["63a breaker", "63amp breaker", "63a ac breaker", "63 amp breaker"]},
    "ac breaker 100a":{"price": 1300,  "unit": "pc",   "aliases": ["100a breaker", "100amp breaker", "100a ac breaker", "100 amp breaker"]},
    # SPD
    "dc spd":         {"price": 780,   "unit": "pc",   "aliases": ["dc surge", "dc surge protector", "dc spd 1200v"]},
    "ac spd 2p":      {"price": 580,   "unit": "pc",   "aliases": ["ac spd 2pole", "ac surge 2p", "2p spd", "2 pole spd"]},
    "ac spd 4p":      {"price": 980,   "unit": "pc",   "aliases": ["ac spd 4pole", "ac surge 4p", "4p spd", "4 pole spd"]},
    # MC4
    "mc4":            {"price": 80,    "unit": "pair", "aliases": ["mc4 connector", "mc4 connectors", "mc-4", "mc 4"]},
    # Battery Breaker
    "battery breaker":{"price": 1700,  "unit": "pc",   "aliases": ["battery cb", "battery circuit breaker", "batt breaker", "battery braker"]},
    # PV Cable
    "pv cable 4mm":   {"price": 70,    "unit": "meter","aliases": ["4mm cable", "4mm wire", "4mm pv cable", "4mm solar wire", "4mm solar cable"]},
    "pv cable 6mm":   {"price": 85,    "unit": "meter","aliases": ["6mm cable", "6mm wire", "6mm pv cable", "6mm solar wire", "6mm solar cable"]},
    # HDPE Conduit
    "hdpe 25mm":      {"price": 170,   "unit": "meter","aliases": ["hdpe", "hdpe conduit", "hdpe pipe", "hdpe 25", "25mm hdpe", "hdpe conduit 25mm", "hdpe pipe 25mm", "hdpe tubing"]},
    # Solar Panels (Talesun)
    # NOTE: generic misspellings (no wattage) default to 620W price
    "solar panel 620w": {"price": 6100, "unit": "pc", "aliases": [
        "620w panel", "620w", "620 watt", "620 watt panel", "620w solar", "620w solar panel",
        "talesun 620w", "talesun 620", "panel 620w", "panel 620", "620",
        # 620W misspellings
        "620w pannel", "620w pannels", "pannels 620w", "pannel 620w",
        "620w panal", "620w panals", "620w panle", "620w panles",
        # Generic panel words (no wattage) — catches misspellings like "pannels", "pannel"
        "panel", "panels", "pannel", "pannels", "panals", "panal",
        "panles", "panle", "solar pannel", "solar pannels", "solar panal",
        "solar panals", "solar panle", "solar panles",
        "solarpanel", "solarpanels", "solor panel", "solor panels",
        "soler panel", "soler panels", "solar pane", "solar panes",
        "solar panels", "solar panel",
    ]},
    "solar panel 585w": {"price": 5750, "unit": "pc", "aliases": [
        "585w panel", "585w", "585 watt", "585 watt panel", "585w solar", "585w solar panel",
        "talesun 585w", "talesun 585", "panel 585w", "panel 585", "585",
        # 585W misspellings
        "585w pannel", "585w pannels", "pannels 585w", "pannel 585w",
        "585w panal", "585w panals", "585w panle", "585w panles",
    ]},
    # GoodWe Inverters
    "goodwe inverter 8kw": {"price": 72000, "unit": "pc", "aliases": [
        "goodwe 8kw", "goodwe 8k", "8kw inverter", "8kw goodwe", "inverter 8kw",
        "8k inverter", "8kw gw", "gw 8kw", "goodwe8kw"
    ]},
    "goodwe inverter 10kw": {"price": 83000, "unit": "pc", "aliases": [
        "goodwe 10kw", "goodwe 10k", "10kw inverter", "10kw goodwe", "inverter 10kw",
        "10k inverter", "10kw gw", "gw 10kw", "goodwe10kw"
    ]},
    "goodwe inverter 12kw": {"price": 92000, "unit": "pc", "aliases": [
        "goodwe 12kw", "goodwe 12k", "12kw inverter", "12kw goodwe", "inverter 12kw",
        "12k inverter", "12kw gw", "gw 12kw", "goodwe12kw"
    ]},
    # GoodWe Battery
    "goodwe battery 280ah": {"price": 118000, "unit": "pc", "aliases": [
        "goodwe battery", "goodwe 280ah", "280ah battery", "battery 280ah",
        "goodwe batt", "gw battery", "gw 280ah", "280ah goodwe",
        "goodwe lithium", "goodwe storage", "goodwe 280"
    ]},
}

# Regex patterns for quantity extraction
import re as _re

# Build a sorted list of (key, data) with longest keys first so specific items
# (e.g. "battery breaker", "ac breaker 100a") are tried before shorter ones.
def _sorted_unit_prices():
    items = list(UNIT_PRICES.items())
    items.sort(key=lambda x: len(x[0]), reverse=True)
    return items

def _resolve_item(raw_name):
    """Match a raw item name to a UNIT_PRICES key. Returns (key, item_data) or (None, None)."""
    raw = raw_name.strip().lower()
    sorted_items = _sorted_unit_prices()

    # 1. Exact key match
    if raw in UNIT_PRICES:
        return raw, UNIT_PRICES[raw]

    # 2. Exact alias match (longest key first)
    for key, data in sorted_items:
        if raw in [a.lower() for a in data["aliases"]]:
            return key, data

    # 3. Contained match: raw fully contained in key/alias OR key/alias fully contained in raw
    #    Longest key wins to avoid short keys swallowing longer ones.
    for key, data in sorted_items:
        if raw == key or key == raw:
            return key, data
        for alias in data["aliases"]:
            al = alias.lower()
            if raw == al:
                return key, data

    # 4. Substring match — only if the candidate key/alias is fully contained in raw
    #    (e.g. "ac breaker 40a" is in "ac breaker 40a connector")
    for key, data in sorted_items:
        if key in raw:
            return key, data
        for alias in data["aliases"]:
            if alias.lower() in raw:
                return key, data

    return None, None

# Tokenise a message into (qty, item_text) pairs.
# Handles: "12pcs railings", "12 pcs railings", "12 pieces railings",
#          "12x railings", "12 railings", "12m 4mm cable"
_QTY_PATTERN = _re.compile(
    r'(\d+)\s*(?:pcs?\.?|pieces?|units?|x|rolls?|meters?|mtrs?|m(?=\s))?' 
    r'\s*([a-z][a-z0-9 \-]*)'
    r'(?=\s*(?:,|\band\b|\bat\b|&|\+|$|\d))',
    _re.IGNORECASE
)

def parse_cart(message):
    """
    Parse a message for quantity+item pairs.
    Returns a list of dicts: [{qty, key, label, unit_price, unit, subtotal}, ...]
    or an empty list if nothing was found.
    Handles: commas, newlines, 'and', 'at', decimal sizes like '2.4m rail'
    """
    text = message.lower()
    # Normalise separators: commas, semicolons, newlines all become ' , '
    text = _re.sub(r'[,;\n\r]+', ' , ', text)
    # English and Filipino separators
    text = _re.sub(r'\b(and|at|&|\+|kasama|pati|at saka|tapos|plus|with|including|kasama na)\b', ' , ', text, flags=_re.IGNORECASE)
    # Remove decimal size prefixes like '2.4m' that precede an item name
    # e.g. '14 pcs 2.4m rail' → '14 pcs rail'  (the alias '2.4m rail' handles matching)
    text = _re.sub(r'\b(\d+\.\d+)\s*m\s+', '', text)

    # Strip leading non-numeric words (e.g. "magkano total", "how much", "bale eto order ko")
    text = _re.sub(r'^[^\d,]+(?=\d)', '', text)
    # Strip trailing Filipino filler words (e.g. "po sana", "po lang", "po", "lang", "nalang")
    text = _re.sub(r'\b(po sana|po lang|po na|po|lang|nalang|na lang|sana|daw|raw)\b\s*$', '', text, flags=_re.IGNORECASE)

    # Split on commas to get individual item segments
    segments = [s.strip() for s in text.split(',') if s.strip()]

    found = []
    seen_keys = set()

    for seg in segments:
        seg = seg.strip()
        # Strip trailing Filipino filler words from each segment
        seg = _re.sub(r'\\b(po sana|po lang|po na|po|lang|nalang|na lang|sana|daw|raw)\\b\\s*$', '', seg, flags=_re.IGNORECASE).strip()

        # Pattern B: item BEFORE qty  → "railings 30pcs", "Lft 100pcs"
        # MUST be checked on the ORIGINAL segment before stripping leading words
        m_after = _re.match(
            r'^([a-z][a-z0-9 \-]*)\s+(\d+)\s*(?:pcs?\.?|pieces?|units?|x|rolls?|meters?|mtrs?|m(?=\s))?\s*$',
            seg, _re.IGNORECASE
        )

        # Strip any leading non-numeric words within a segment (for Pattern A)
        seg_stripped = _re.sub(r'^[^\d]+(?=\d)', '', seg)

        # Pattern A: qty BEFORE item  → "20pcs 620W", "30 railings"
        m_before = _re.match(
            r'^(\d+)\s*(?:pcs?\.?|pieces?|units?|x|rolls?|meters?|mtrs?|m(?=\s))?\s*(.+)$',
            seg_stripped, _re.IGNORECASE
        )

        candidates = []
        if m_before:
            candidates.append((int(m_before.group(1)), m_before.group(2).strip()))
        if m_after:
            candidates.append((int(m_after.group(2)), m_after.group(1).strip()))

        for qty, raw_name in candidates:
            key, data = _resolve_item(raw_name)
            if key and key not in seen_keys:
                seen_keys.add(key)
                found.append({
                    "qty": qty,
                    "key": key,
                    "label": key.title(),
                    "unit_price": data["price"],
                    "unit": data["unit"],
                    "subtotal": qty * data["price"],
                })
                break  # stop after first successful match for this segment
    return found

def format_cart_response(cart, language):
    """Build the itemized total message from a parsed cart."""
    grand_total = sum(item["subtotal"] for item in cart)
    lines = []
    for item in cart:
        lines.append(
            f"- {item['qty']} {item['unit']}(s) × {item['label']}: "
            f"₱{item['unit_price']:,}/{item['unit']} = "
            f"₱{item['subtotal']:,}"
        )
    items_block = "\n".join(lines)

    if language == "tl":
        return (
            f"🧮 **Listahan ng Presyo:**\n\n"
            f"{items_block}\n\n"
            f"💰 **KABUUANG HALAGA: ₱{grand_total:,}**\n\n"
            f"📌 *Ang presyo ay para sa materyales lamang. Hindi pa kasama ang delivery at labor.*\n"
            f"Makipag-ugnayan sa amin para sa opisyal na quotation! 💚"
        )
    else:
        return (
            f"🧮 **Price Breakdown:**\n\n"
            f"{items_block}\n\n"
            f"💰 **TOTAL: ₱{grand_total:,}**\n\n"
            f"📌 *Prices are for materials only. Delivery and labor charges not yet included.*\n"
            f"Contact us for an official quotation! 💚"
        )

# ─────────────────────────────────────────────────────────────────────────────

def detect_language(text):
    """Detect if message is in Tagalog or English"""
    tagalog_words = ["ang", "sa", "ng", "ko", "mo", "nyo", "kami", "tayo", "sila", "po", "ba", "kayo", "magkano", "saan", "paano", "ano", "ito", "yan", "dito", "doon", "nandito", "nandoon"]
    
    words = text.lower().split()
    tagalog_count = sum(1 for word in words if any(tl_word in word for tl_word in tagalog_words))
    
    # If more than 30% of words are Tagalog indicators, it's Tagalog
    if len(words) > 0 and tagalog_count / len(words) > 0.3:
        return "tl"
    return "en"

def save_user_language(user_id, language):
    """Save user's language preference"""
    try:
        with db_lock:
            conn = sqlite3.connect('/tmp/ultiphoton_chatbot.db')
            cursor = conn.cursor()
            cursor.execute('INSERT OR REPLACE INTO user_preferences (user_id, language) VALUES (?, ?)', (user_id, language))
            conn.commit()
            conn.close()
    except Exception as e:
        print(f"❌ Error saving language: {str(e)}")
        sys.stdout.flush()

def get_user_language(user_id):
    """Get user's saved language preference"""
    try:
        with db_lock:
            conn = sqlite3.connect('/tmp/ultiphoton_chatbot.db')
            cursor = conn.cursor()
            cursor.execute('SELECT language FROM user_preferences WHERE user_id = ?', (user_id,))
            result = cursor.fetchone()
            conn.close()
            return result[0] if result else "auto"
    except:
        return "auto"

def save_pending_panel_qty(user_id, qty):
    """Save pending panel quantity while waiting for wattage clarification."""
    try:
        with db_lock:
            conn = sqlite3.connect('/tmp/ultiphoton_chatbot.db')
            cursor = conn.cursor()
            cursor.execute(
                'INSERT OR REPLACE INTO user_preferences (user_id, pending_panel_qty) '
                'VALUES (?, ?) ON CONFLICT(user_id) DO UPDATE SET pending_panel_qty=excluded.pending_panel_qty',
                (user_id, qty)
            )
            conn.commit()
            conn.close()
    except Exception as e:
        print(f"❌ Error saving pending_panel_qty: {e}")

def get_pending_panel_qty(user_id):
    """Get pending panel quantity for wattage clarification."""
    try:
        with db_lock:
            conn = sqlite3.connect('/tmp/ultiphoton_chatbot.db')
            cursor = conn.cursor()
            cursor.execute('SELECT pending_panel_qty FROM user_preferences WHERE user_id = ?', (user_id,))
            result = cursor.fetchone()
            conn.close()
            return result[0] if result and result[0] is not None else None
    except:
        return None

def clear_pending_panel_qty(user_id):
    """Clear pending panel quantity after wattage is resolved."""
    try:
        with db_lock:
            conn = sqlite3.connect('/tmp/ultiphoton_chatbot.db')
            cursor = conn.cursor()
            cursor.execute(
                'UPDATE user_preferences SET pending_panel_qty = NULL WHERE user_id = ?',
                (user_id,)
            )
            conn.commit()
            conn.close()
    except Exception as e:
        print(f"❌ Error clearing pending_panel_qty: {e}")

def save_pending_wattage(user_id, wattage):
    """Save pending wattage while waiting for price tier choice."""
    try:
        with db_lock:
            conn = sqlite3.connect('/tmp/ultiphoton_chatbot.db')
            cursor = conn.cursor()
            cursor.execute(
                'UPDATE user_preferences SET pending_wattage = ? WHERE user_id = ?',
                (wattage, user_id)
            )
            conn.commit()
            conn.close()
    except Exception as e:
        print(f"❌ Error saving pending_wattage: {e}")

def get_pending_wattage(user_id):
    """Get pending wattage for price tier clarification."""
    try:
        with db_lock:
            conn = sqlite3.connect('/tmp/ultiphoton_chatbot.db')
            cursor = conn.cursor()
            cursor.execute('SELECT pending_wattage FROM user_preferences WHERE user_id = ?', (user_id,))
            result = cursor.fetchone()
            conn.close()
            return result[0] if result and result[0] is not None else None
    except:
        return None

def clear_pending_wattage(user_id):
    """Clear pending wattage after price tier is resolved."""
    try:
        with db_lock:
            conn = sqlite3.connect('/tmp/ultiphoton_chatbot.db')
            cursor = conn.cursor()
            cursor.execute(
                'UPDATE user_preferences SET pending_wattage = NULL WHERE user_id = ?',
                (user_id,)
            )
            conn.commit()
            conn.close()
    except Exception as e:
        print(f"❌ Error clearing pending_wattage: {e}")

# ── Persistent greeting store (survives server restarts) ─────────────────────
# Stored as a JSON file alongside app.py so Render keeps it between deploys.
# Format: { "<user_id>": "YYYY-MM-DD", ... }
import json as _json

_GREETING_FILE = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "greeting_dates.json")
_greeting_lock = Lock()

def _load_greeting_store():
    try:
        with open(_GREETING_FILE, "r") as f:
            return _json.load(f)
    except Exception:
        return {}

def _save_greeting_store(store):
    try:
        with open(_GREETING_FILE, "w") as f:
            _json.dump(store, f)
    except Exception as e:
        print(f"❌ Error saving greeting store: {e}")
        sys.stdout.flush()

def _ph_today():
    """Return today's date string in Philippine Time (UTC+8), resets at midnight PH."""
    from datetime import datetime, timezone, timedelta
    ph_tz = timezone(timedelta(hours=8))
    return datetime.now(ph_tz).strftime("%Y-%m-%d")

def should_send_greeting(user_id):
    """Return True if the greeting has NOT been sent today (PH Time) for this user."""
    try:
        today = _ph_today()
        with _greeting_lock:
            store = _load_greeting_store()
        return store.get(str(user_id)) != today
    except Exception:
        return True

def mark_greeting_sent(user_id):
    """Record that the greeting was sent today (PH Time) for this user."""
    try:
        today = _ph_today()
        with _greeting_lock:
            store = _load_greeting_store()
            store[str(user_id)] = today
            _save_greeting_store(store)
        print(f"✅ Greeting marked for {user_id} on {today}")
        sys.stdout.flush()
    except Exception as e:
        print(f"❌ Error marking greeting sent: {e}")
        sys.stdout.flush()

# Keep old names as aliases so nothing else breaks
def is_first_message(user_id):
    return should_send_greeting(user_id)

def mark_first_message_sent(user_id):
    mark_greeting_sent(user_id)

def save_conversation(user_id, message, response, language, faq_matched):
    """Save conversation to database for analytics"""
    try:
        with db_lock:
            conn = sqlite3.connect('/tmp/ultiphoton_chatbot.db')
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO conversations (user_id, message, response, language, faq_matched)
                VALUES (?, ?, ?, ?, ?)
            ''', (user_id, message, response, language, faq_matched))
            conn.commit()
            conn.close()
    except Exception as e:
        print(f"❌ Error saving conversation: {str(e)}")
        sys.stdout.flush()

def log_analytics(user_id, faq_key, keyword):
    """Log analytics for FAQ usage"""
    try:
        with db_lock:
            conn = sqlite3.connect('/tmp/ultiphoton_chatbot.db')
            cursor = conn.cursor()
            cursor.execute('INSERT INTO analytics (user_id, faq_key, keyword) VALUES (?, ?, ?)', (user_id, faq_key, keyword))
            conn.commit()
            conn.close()
    except Exception as e:
        print(f"❌ Error logging analytics: {str(e)}")
        sys.stdout.flush()

def get_analytics_summary():
    """Get analytics summary for dashboard"""
    try:
        with db_lock:
            conn = sqlite3.connect('/tmp/ultiphoton_chatbot.db')
            cursor = conn.cursor()
            
            # Total conversations
            cursor.execute('SELECT COUNT(*) FROM conversations')
            total_conversations = cursor.fetchone()[0]
            
            # FAQ matches
            cursor.execute('SELECT COUNT(*) FROM conversations WHERE faq_matched = 1')
            faq_matches = cursor.fetchone()[0]
            
            # Most popular FAQs
            cursor.execute('''
                SELECT faq_key, COUNT(*) as count FROM analytics 
                GROUP BY faq_key ORDER BY count DESC LIMIT 5
            ''')
            popular_faqs = cursor.fetchall()
            
            # Language distribution
            cursor.execute('''
                SELECT language, COUNT(*) as count FROM conversations 
                GROUP BY language
            ''')
            language_dist = cursor.fetchall()
            
            conn.close()
            
            return {
                "total_conversations": total_conversations,
                "faq_matches": faq_matches,
                "popular_faqs": popular_faqs,
                "language_distribution": language_dist
            }
    except Exception as e:
        print(f"❌ Error getting analytics: {str(e)}")
        sys.stdout.flush()
        return {}

def is_business_hours():
    """Check if current time is within business hours (Philippine Time)"""
    from datetime import datetime, timezone, timedelta
    
    # Philippine Time is UTC+8
    ph_tz = timezone(timedelta(hours=8))
    current_time = datetime.now(ph_tz)
    current_hour = current_time.hour
    current_day = current_time.weekday()
    
    return (BUSINESS_HOURS["start"] <= current_hour < BUSINESS_HOURS["end"] and 
            current_day in BUSINESS_HOURS["days"])

def get_greeting_message(language):
    """Get auto-greeting message"""
    if language == "tl":
        return """👋 **Maligayang pagdating sa Ultiphoton Solar Power OPC!** ☀️

Kami ay nandito upang sagutin ang lahat ng inyong mga tanong tungkol sa solar panels at renewable energy solutions.

Paano namin kayo matutulungan ngayong araw? 🌞

Maaari kayong magtanong tungkol sa:
• 💰 Presyo ng Solar Panels
• 📍 Aming Lokasyon
• 💳 Paraan ng Pagbabayad
• 🔧 Accessories at Mounting
• ⚡ Inverter Brands
• 📦 Delivery at Installation

Salamat sa inyong interes! 💚"""
    else:
        return """👋 **Welcome to Ultiphoton Solar Power OPC!** ☀️

We're here to answer all your questions about solar panels and renewable energy solutions.

How can we help you today? 🌞

You can ask about:
• 💰 Solar Panel Pricing
• 📍 Our Locations
• 💳 Payment Methods
• 🔧 Accessories & Mounting
• ⚡ Inverter Brands
• 📦 Delivery & Installation

Thank you for your interest! 💚"""

def get_business_hours_message(language):
    """Legacy function kept for compatibility — now returns after-hours note."""
    return get_after_hours_note(language)

def get_after_hours_note(language):
    """Brief after-hours note appended to every response outside business hours."""
    if language == "tl":
        return ("🌙 *Paunawa:* Kami ay kasalukuyang nasa labas ng aming business hours "
                "(Mon–Sun, 8AM–6PM PH Time). Nakatanggap ka ng automated na sagot. "
                "Ang aming team ay mag-follow up sa iyo sa susunod na araw ng trabaho! 💚")
    else:
        return ("🌙 *Note:* We are currently outside our business hours "
                "(Mon–Sun, 8AM–6PM PH Time). You have received an automated reply. "
                "Our team will follow up with you on the next business day! 💚")

# Per-item accessory FAQ keys — checked BEFORE the full accessories list
# Order matters: more specific items first (e.g. dc_breaker before ac_breaker)
PER_ITEM_ACCESSORY_KEYS = [
    "rail_splicer", "railing",   # rail_splicer MUST come before railing (substring issue)
    "l_foot", "mid_clamp", "end_clamp",
    "grounding_lug", "dc_breaker", "ac_breaker", "dc_spd", "ac_spd",
    "mc4", "battery_breaker",
    "pv_cable", "hdpe_price",
    # Installation estimates — longer kW numbers MUST come before shorter ones
    # (e.g. "15kw" must be checked before "5kw" to avoid substring match)
    "install_15kw", "install_12kw", "install_10kw",
    "install_8kw", "install_6kw", "install_5kw", "install_3kw",
    "install_general",
]

# Keywords that trigger the FULL accessories list (only if no per-item match)
ACCESSORY_FULL_LIST_KEYWORDS = {
    "mounting", "mountings", "soeasy", "chint", "chyt", "surge protection",
    "accessories", "materials", "pv mounting"
}

# Keywords that are specific to solar panels — must be checked BEFORE generic pricing keywords
SOLAR_SPECIFIC_KEYWORDS = {
    "solar panel", "solar panels", "talesun", "585w", "620w", "panel price", "solar price",
    "bifacial", "photovoltaic", "pv panel"
}

def find_matching_faq(user_message):
    """Find matching FAQ with priority:
    1. Per-item accessory keywords (railing, mc4, dc breaker, etc.) — most specific
    2. Full accessories list (generic accessory queries)
    3. Solar-panel-specific keywords
    4. All other FAQs via their keyword lists
    5. Fallback to accessories / solar_panel_price via their generic keywords
    """
    message_lower = user_message.lower()

    # --- PASS 1: Per-item accessory keywords (most specific) ---
    for item_key in PER_ITEM_ACCESSORY_KEYS:
        faq_data = FAQS.get(item_key)
        if faq_data:
            for keyword in faq_data["keywords"]:
                if keyword.lower() in message_lower:
                    return item_key, faq_data

    # --- PASS 2: Full accessories list (generic accessory queries) ---
    for kw in ACCESSORY_FULL_LIST_KEYWORDS:
        if kw in message_lower:
            return "accessories", FAQS["accessories"]

    # --- PASS 3: Check solar-panel-specific keywords ---
    for kw in SOLAR_SPECIFIC_KEYWORDS:
        if kw in message_lower:
            return "solar_panel_price", FAQS["solar_panel_price"]

    # --- PASS 4: Normal FAQ loop (skip per-item, accessories, solar_panel_price) ---
    SKIP_IN_PASS4 = set(PER_ITEM_ACCESSORY_KEYS) | {"accessories", "solar_panel_price"}
    for faq_key, faq_data in FAQS.items():
        if faq_key in SKIP_IN_PASS4:
            continue
        for keyword in faq_data["keywords"]:
            if keyword.lower() in message_lower:
                return faq_key, faq_data

    # --- PASS 5: Fall back to accessories / solar_panel_price via their generic keywords ---
    for faq_key in ("accessories", "solar_panel_price"):
        faq_data = FAQS[faq_key]
        for keyword in faq_data["keywords"]:
            if keyword.lower() in message_lower:
                return faq_key, faq_data

    return None, None

def get_faq_answer(faq_data, language, faq_key=None):
    """Get FAQ answer - uses live Excel data for pricing FAQs, static text for others."""
    # For pricing FAQs, always build answer from the Excel file
    if faq_key == "solar_panel_price":
        excel_answer = build_solar_panel_answer(language)
        if excel_answer:
            return excel_answer
    elif faq_key == "accessories":
        excel_answer = build_accessories_answer(language)
        if excel_answer:
            return excel_answer
    elif faq_key in ("full_pricelist", "products_and_price"):
        excel_answer = build_pricelist_answer(language)
        if excel_answer:
            return excel_answer

    # Fallback to static FAQ text for non-pricing FAQs or if Excel unavailable
    if language == "tl" and "answer_tl" in faq_data:
        return faq_data["answer_tl"]
    elif "answer_en" in faq_data:
        return faq_data["answer_en"]
    else:
        return faq_data.get("answer", "")

# ─────────────────────────────────────────────────────────────────────────────
# WATTAGE CLARIFICATION FLOW
# When customer asks about N panels without specifying 620W or 585W,
# bot asks which wattage, then calculates total package price.
# ─────────────────────────────────────────────────────────────────────────────
import re as _re_watt

# Keywords that indicate a panel quantity inquiry (no wattage specified)
_PANEL_QTY_INQUIRY_KEYWORDS = [
    "solar panel", "solar panels", "solar pannel", "solar pannels",
    "solor panel", "soler panel", "pv panel",
    "pannel", "pannels", "panal", "panals", "panle", "panles",
    "panel", "panels",
]
# Keywords that indicate wattage IS specified (skip clarification)
_WATTAGE_SPECIFIED_PATTERNS = [
    r'\b620\b', r'\b585\b', r'\b620w\b', r'\b585w\b',
    r'620\s*watt', r'585\s*watt',
]
# Keywords that indicate customer is replying with wattage choice
_WATTAGE_REPLY_620 = [
    "620", "620w", "620 watt", "620watts", "six twenty",
    "620w po", "620 po", "yung 620", "ang 620",
]
_WATTAGE_REPLY_585 = [
    "585", "585w", "585 watt", "585watts", "five eighty five",
    "585w po", "585 po", "yung 585", "ang 585",
]
# Keywords that indicate customer is replying with price tier choice
_PRICE_TIER_RETAIL = [
    "retail", "retail price", "retail po", "regular", "regular price",
    "normal", "normal price", "retail lang", "retail na",
    "1", "option 1", "una", "first",
]
_PRICE_TIER_INSTALLER = [
    "installer", "installer price", "installer po", "installer lang",
    "installer na", "install", "instaler", "instaler price",
    "2", "option 2", "pangalawa", "second",
]
# Threshold for offering installer price
_INSTALLER_THRESHOLD = 10

def get_recent_wattage_question(user_id):
    """Recover pending panel qty from recent conversation history.
    Returns panel_count if the last bot message was a wattage question, else None."""
    import re as _re
    try:
        with db_lock:
            conn = sqlite3.connect('/tmp/ultiphoton_chatbot.db')
            cursor = conn.cursor()
            # Get the last bot response for this user within the last 10 minutes
            cursor.execute(
                """SELECT response FROM conversations
                   WHERE user_id = ? AND faq_matched = 1
                   AND timestamp >= datetime('now', '-10 minutes')
                   ORDER BY timestamp DESC LIMIT 1""",
                (user_id,)
            )
            result = cursor.fetchone()
            conn.close()
            if result:
                resp = result[0]
                # Check if the last bot message was a wattage clarification question
                if ('620' in resp or '585' in resp) and ('wattage' in resp.lower() or 'reply' in resp.lower() or 'i-reply' in resp.lower() or 'anong' in resp.lower()):
                    # Extract panel count from the response text
                    nums = _re.findall(r'(\d+)\s*(?:pcs?|panels?|solar panels?)', resp)
                    if nums:
                        qty = int(nums[0])
                        if 1 <= qty <= 500:
                            return qty
            return None
    except:
        return None

def detect_panel_qty_no_wattage(user_message):
    """Detect if message asks about N panels WITHOUT specifying wattage.
    Returns panel_count (int) if ambiguous, else None."""
    msg = user_message.lower()
    # Must mention panels
    if not any(kw in msg for kw in _PANEL_QTY_INQUIRY_KEYWORDS):
        return None
    # Must NOT already specify wattage
    for pat in _WATTAGE_SPECIFIED_PATTERNS:
        if _re_watt.search(pat, msg):
            return None
    # Must contain a number (panel count) — also handles "4pcs", "10pc"
    numbers = _re_watt.findall(r'(\d+)\s*(?:pcs?\.?|pieces?|units?|piraso)?', msg)
    candidates = [int(n) for n in numbers if 1 <= int(n) <= 500]
    if not candidates:
        return None
    return candidates[0]

def detect_wattage_reply(user_message):
    """Detect if user is replying with a wattage choice (620 or 585).
    Returns 620, 585, or None."""
    msg = user_message.lower().strip()
    for kw in _WATTAGE_REPLY_620:
        if kw in msg:
            return 620
    for kw in _WATTAGE_REPLY_585:
        if kw in msg:
            return 585
    return None

def detect_price_tier_reply(user_message):
    """Detect if user is replying with a price tier choice (retail or installer).
    Returns 'retail', 'installer', or None."""
    msg = user_message.lower().strip()
    # Check installer first (more specific)
    for kw in _PRICE_TIER_INSTALLER:
        if kw in msg:
            return 'installer'
    for kw in _PRICE_TIER_RETAIL:
        if kw in msg:
            return 'retail'
    return None

def ask_price_tier(panel_count, wattage, language):
    """Ask customer whether they want retail or installer pricing (for 10+ panels)."""
    retail_price    = 6100 if wattage == 620 else 5750
    installer_price = 5850 if wattage == 620 else 5650
    panel_label     = f"Talesun {wattage}W"
    if language == "tl":
        return (
            f"☀️ {panel_count} pcs {panel_label} — Anong presyo po ang gusto ninyo?\n\n"
            f"🔵 **1. Retail Price** — ₱{retail_price:,}/pc\n"
            f"🟢 **2. Installer Price** — ₱{installer_price:,}/pc *(para sa installers/bulk orders)*\n\n"
            f"I-reply lang po ng **retail** o **installer** para makuha ang kabuuang presyo! 😊"
        )
    else:
        return (
            f"☀️ {panel_count} pcs {panel_label} — Which price tier do you prefer?\n\n"
            f"🔵 **1. Retail Price** — ₱{retail_price:,}/pc\n"
            f"🟢 **2. Installer Price** — ₱{installer_price:,}/pc *(for installers/bulk orders)*\n\n"
            f"Just reply **retail** or **installer** to get the full package price! 😊"
        )

def format_panel_package_response(panel_count, wattage, language, price_tier='retail'):
    """Calculate and format the full panel package price breakdown."""
    import math
    if price_tier == 'installer':
        price_per_panel = 5850 if wattage == 620 else 5650
        tier_label = "Installer Price"
    else:
        price_per_panel = 6100 if wattage == 620 else 5750
        tier_label = "Retail Price"
    panel_label = f"Talesun {wattage}W"

    # Panel subtotal
    panel_total = panel_count * price_per_panel

    # Mounting hardware
    railings     = math.ceil(panel_count * 1.2)
    l_foot       = panel_count * 3
    mid_clamp    = panel_count * 2
    end_clamp    = panel_count * 2
    rail_splicer = panel_count
    pv_lug       = panel_count

    t_railing    = railings * 600
    t_l_foot     = l_foot * 95
    t_mid_clamp  = mid_clamp * 85
    t_end_clamp  = end_clamp * 85
    t_splicer    = rail_splicer * 85
    t_lug        = pv_lug * 70
    mounting_total = t_railing + t_l_foot + t_mid_clamp + t_end_clamp + t_splicer + t_lug
    grand_total  = panel_total + mounting_total

    if language == "tl":
        return (
            f"☀️ **Package para sa {panel_count} pcs {panel_label} ({tier_label}):**\n\n"
            f"📊 **Solar Panels:**\n"
            f"- {panel_count} pcs × {panel_label}: ₱{price_per_panel:,}/pc = ₱{panel_total:,}\n\n"
            f"🔧 **Mounting Hardware:**\n"
            f"- Railing 2.4m: {railings} pcs × ₱600 = ₱{t_railing:,}\n"
            f"- L-Foot: {l_foot} pcs × ₱95 = ₱{t_l_foot:,}\n"
            f"- Mid Clamp: {mid_clamp} pcs × ₱85 = ₱{t_mid_clamp:,}\n"
            f"- End Clamp: {end_clamp} pcs × ₱85 = ₱{t_end_clamp:,}\n"
            f"- Rail Splicer: {rail_splicer} pcs × ₱85 = ₱{t_splicer:,}\n"
            f"- PV Grounding Lug: {pv_lug} pcs × ₱70 = ₱{t_lug:,}\n\n"
            f"💰 **KABUUANG HALAGA: ₱{grand_total:,}**\n"
            f"  *(Panels: ₱{panel_total:,} + Mounting: ₱{mounting_total:,})*\n\n"
            f"📌 *Para sa materyales lamang. Hindi pa kasama ang delivery at labor.*\n"
            f"Makipag-ugnayan sa amin para sa opisyal na quotation! 💚"
        )
    else:
        return (
            f"☀️ **Package for {panel_count} pcs {panel_label} ({tier_label}):**\n\n"
            f"📊 **Solar Panels:**\n"
            f"- {panel_count} pcs × {panel_label}: ₱{price_per_panel:,}/pc = ₱{panel_total:,}\n\n"
            f"🔧 **Mounting Hardware:**\n"
            f"- Railing 2.4m: {railings} pcs × ₱600 = ₱{t_railing:,}\n"
            f"- L-Foot: {l_foot} pcs × ₱95 = ₱{t_l_foot:,}\n"
            f"- Mid Clamp: {mid_clamp} pcs × ₱85 = ₱{t_mid_clamp:,}\n"
            f"- End Clamp: {end_clamp} pcs × ₱85 = ₱{t_end_clamp:,}\n"
            f"- Rail Splicer: {rail_splicer} pcs × ₱85 = ₱{t_splicer:,}\n"
            f"- PV Grounding Lug: {pv_lug} pcs × ₱70 = ₱{t_lug:,}\n\n"
            f"💰 **TOTAL: ₱{grand_total:,}**\n"
            f"  *(Panels: ₱{panel_total:,} + Mounting: ₱{mounting_total:,})*\n\n"
            f"📌 *Prices are for materials only. Delivery and labor not yet included.*\n"
            f"Contact us for an official quotation! 💚"
        )

def ask_wattage_question(panel_count, language):
    """Generate the wattage clarification question."""
    if language == "tl":
        return (
            f"☀️ Salamat po sa inyong interes sa {panel_count} solar panels!\n\n"
            f"👉 Anong wattage po ang gusto ninyo?\n\n"
            f"🔵 **620W** — ₱6,100/pc (mas mataas na output)\n"
            f"🟡 **585W** — ₱5,750/pc (mas ekonomikal)\n\n"
            f"I-reply lang po ng **620** o **585** para makuha ang kabuuang presyo kasama ang mounting hardware! 😊"
        )
    else:
        return (
            f"☀️ Thank you for your interest in {panel_count} solar panels!\n\n"
            f"👉 Which wattage do you prefer?\n\n"
            f"🔵 **620W** — ₱6,100/pc (higher output)\n"
            f"🟡 **585W** — ₱5,750/pc (more economical)\n\n"
            f"Just reply **620** or **585** and I'll calculate the full package price including mounting hardware! 😊"
        )

# ─────────────────────────────────────────────────────────────────────────────
# MOUNTING HARDWARE CALCULATOR
# Triggered when customer asks how many accessories they need for N panels
# ─────────────────────────────────────────────────────────────────────────────
import re as _re_hw
import math

_HW_TRIGGER_KEYWORDS = [
    # Filipino
    "ilang l-foot", "ilang l foot", "ilang lfoot", "ilang lft",
    "ilang mid clamp", "ilang midclamp",
    "ilang end clamp", "ilang endclamp",
    "ilang rail splicer", "ilang splicer",
    "ilang grounding lug", "ilang lug",
    "ilang railing", "ilang railings",
    "ilang piraso ang kailangan", "ilang piraso ang need",
    "ilang kailangan", "ilang need ko", "ilang po kailangan",
    "ilang accessories", "ilang mounting",
    "ilang po ang kailangan", "ilang po ang need",
    # English
    "how many l-foot", "how many l foot", "how many lfoot",
    "how many mid clamp", "how many end clamp",
    "how many rail splicer", "how many splicer",
    "how many grounding lug", "how many lug",
    "how many railing", "how many railings",
    "how many accessories", "how many mounting",
    "how many do i need", "how many pieces do i need",
    "accessories needed", "mounting needed",
    "accessories for", "mounting for",
    "pieces needed", "pcs needed",
]

_HW_PANEL_PATTERN = _re_hw.compile(
    r'(\d+)\s*(?:pcs?\.?|pieces?|panels?|pannels?|solar panels?|units?)?\s*'
    r'(?:panel|panels|pannels|solar|pv)?',
    _re_hw.IGNORECASE
)

def detect_hardware_calc(user_message):
    """Detect if user is asking how many mounting accessories they need for N panels.
    Returns panel_count (int) if detected, else None."""
    msg = user_message.lower()
    # Must contain a trigger keyword
    if not any(kw in msg for kw in _HW_TRIGGER_KEYWORDS):
        return None
    # Extract panel count from message
    numbers = _re_hw.findall(r'\b(\d+)\b', msg)
    if not numbers:
        return None
    # Take the first/largest number as panel count (filter out small noise numbers)
    candidates = [int(n) for n in numbers if 1 <= int(n) <= 500]
    if not candidates:
        return None
    return candidates[0]

def format_hardware_calc_response(panel_count, language):
    """Calculate and format mounting hardware quantities and total price."""
    import math
    # Formulas
    railings    = math.ceil(panel_count * 1.2)
    l_foot      = panel_count * 3
    mid_clamp   = panel_count * 2
    end_clamp   = panel_count * 2
    rail_splicer= panel_count * 1
    pv_lug      = panel_count * 1

    # Prices from UNIT_PRICES
    p_railing   = 600
    p_l_foot    = 95
    p_mid_clamp = 85
    p_end_clamp = 85
    p_splicer   = 85
    p_lug       = 70

    # Totals
    t_railing   = railings    * p_railing
    t_l_foot    = l_foot      * p_l_foot
    t_mid_clamp = mid_clamp   * p_mid_clamp
    t_end_clamp = end_clamp   * p_end_clamp
    t_splicer   = rail_splicer* p_splicer
    t_lug       = pv_lug      * p_lug
    grand_total = t_railing + t_l_foot + t_mid_clamp + t_end_clamp + t_splicer + t_lug

    if language == "tl":
        return (
            f"🔧 **Mounting Hardware para sa {panel_count} Solar Panels:**\n\n"
            f"| Materyales | Qty | Presyo/pc | Subtotal |\n"
            f"|---|---|---|---|\n"
            f"| Aluminum Railing 2.4m | {railings} pcs | ₱{p_railing:,} | ₱{t_railing:,} |\n"
            f"| L-Foot | {l_foot} pcs | ₱{p_l_foot:,} | ₱{t_l_foot:,} |\n"
            f"| Mid Clamp | {mid_clamp} pcs | ₱{p_mid_clamp:,} | ₱{t_mid_clamp:,} |\n"
            f"| End Clamp | {end_clamp} pcs | ₱{p_end_clamp:,} | ₱{t_end_clamp:,} |\n"
            f"| Rail Splicer | {rail_splicer} pcs | ₱{p_splicer:,} | ₱{t_splicer:,} |\n"
            f"| PV Grounding Lug | {pv_lug} pcs | ₱{p_lug:,} | ₱{t_lug:,} |\n\n"
            f"💰 **KABUUANG HALAGA: ₱{grand_total:,}**\n\n"
            f"📐 *Formula:*\n"
            f"- Railing = panels × 1.2 (rounded up)\n"
            f"- L-Foot = panels × 3\n"
            f"- Mid/End Clamp = panels × 2\n"
            f"- Rail Splicer / PV Lug = panels × 1\n\n"
            f"📌 *Para sa materyales lamang. Hindi pa kasama ang delivery at labor.*\n"
            f"Makipag-ugnayan sa amin para sa opisyal na quotation! 💚"
        )
    else:
        return (
            f"🔧 **Mounting Hardware for {panel_count} Solar Panels:**\n\n"
            f"| Material | Qty | Price/pc | Subtotal |\n"
            f"|---|---|---|---|\n"
            f"| Aluminum Railing 2.4m | {railings} pcs | ₱{p_railing:,} | ₱{t_railing:,} |\n"
            f"| L-Foot | {l_foot} pcs | ₱{p_l_foot:,} | ₱{t_l_foot:,} |\n"
            f"| Mid Clamp | {mid_clamp} pcs | ₱{p_mid_clamp:,} | ₱{t_mid_clamp:,} |\n"
            f"| End Clamp | {end_clamp} pcs | ₱{p_end_clamp:,} | ₱{t_end_clamp:,} |\n"
            f"| Rail Splicer | {rail_splicer} pcs | ₱{p_splicer:,} | ₱{t_splicer:,} |\n"
            f"| PV Grounding Lug | {pv_lug} pcs | ₱{p_lug:,} | ₱{t_lug:,} |\n\n"
            f"💰 **TOTAL: ₱{grand_total:,}**\n\n"
            f"📐 *Formula used:*\n"
            f"- Railing = panels × 1.2 (rounded up)\n"
            f"- L-Foot = panels × 3\n"
            f"- Mid/End Clamp = panels × 2\n"
            f"- Rail Splicer / PV Lug = panels × 1\n\n"
            f"📌 *Prices are for materials only. Delivery and labor not yet included.*\n"
            f"Contact us for an official quotation! 💚"
        )

# ─────────────────────────────────────────────────────────────────────────────
def get_quick_reply_buttons(language):
    """Get quick reply buttons for common questions"""
    if language == "tl":
        buttons = [
            {"title": "💰 Presyo", "payload": "presyo"},
            {"title": "📍 Lokasyon", "payload": "lokasyon"},
            {"title": "💳 Pagbabayad", "payload": "pagbabayad"},
            {"title": "⚡ Inverter", "payload": "inverter"},
            {"title": "🔧 Accessories", "payload": "accessories"},
            {"title": "📞 Makipag-ugnayan", "payload": "contact"}
        ]
    else:
        buttons = [
            {"title": "💰 Pricing", "payload": "pricing"},
            {"title": "📍 Location", "payload": "location"},
            {"title": "💳 Payment", "payload": "payment"},
            {"title": "⚡ Inverter", "payload": "inverter"},
            {"title": "🔧 Accessories", "payload": "accessories"},
            {"title": "📞 Contact", "payload": "contact"}
        ]
    return buttons

def get_ai_response(user_message, language):
    """Get AI response from OpenAI with FAQ context"""
    try:
        print(f"🤖 Processing: {user_message[:50]}... (Language: {language})")
        sys.stdout.flush()

        # ── Lalamove Delivery Estimate: detect customer address messages ──────────
        lalamove_response = estimate_lalamove_from_message(user_message, language)
        if lalamove_response:
            print(f"🚚 Lalamove estimate generated")
            sys.stdout.flush()
            return lalamove_response, True, "lalamove_estimate"

        # ── Mounting Hardware Calculator: detect "ilang kailangan" type queries ──
        hw_panel_count = detect_hardware_calc(user_message)
        if hw_panel_count:
            print(f"🔧 Hardware calc triggered: {hw_panel_count} panels")
            sys.stdout.flush()
            return format_hardware_calc_response(hw_panel_count, language), True, "hardware_calculator"

        # ── Price Calculator: check for quantity+item pairs first ──────────────
        cart = parse_cart(user_message)
        if cart:
            print(f"🧮 Cart detected: {[i['key'] for i in cart]}")
            sys.stdout.flush()
            return format_cart_response(cart, language), True, "price_calculator"

        # Check for FAQ match first
        faq_key, faq_data = find_matching_faq(user_message)
        if faq_key and faq_data:
            print(f"✅ FAQ Match Found: {faq_key}")
            sys.stdout.flush()
            return get_faq_answer(faq_data, language, faq_key=faq_key), True, faq_key
        
        # If no FAQ match, use AI to generate response
        print(f"🤖 Using AI to generate response...")
        sys.stdout.flush()
        
        headers = {
            "Authorization": f"Bearer {OPENAI_API_KEY}",
            "Content-Type": "application/json"
        }
        
        system_message = """You are a helpful AI assistant for Ultiphoton Solar Power OPC, a solar panel company in the Philippines.

Company Information:
- Main Office: Filinvest, Muntilupa City
- Branch: Batangas
- Warehouse: Cainta
- Products: Talesun Solar Panels (585W, 620W), Inverters (Deye, Solis, GoodWe, SRNE, Sigenergy)
- Services: Installation, Maintenance, Consultation
- Delivery: COD available in Batangas, Laguna, Quezon, South Luzon
- Website: https://ultiphotonsolarpoweropc.com/

Guidelines:
1. Be friendly, professional, and helpful
2. Keep responses concise (under 100 words)
3. If you don't know specific details, suggest they contact the company
4. Always mention "Feel free to contact us!" at the end
5. Use emojis to make responses friendly ☀️⚡💚
6. IMPORTANT: NEVER quote any prices or peso amounts. You do not have access to the official price list.
   If asked about pricing, ALWAYS say: 'For our official price list, please ask: What is the price list?' or direct them to message the page.
7. All pricing information is handled by a separate FAQ system - do not attempt to answer price questions yourself."""
        
        if language == "tl":
            system_message += "\n8. Respond in Tagalog/Filipino"
        else:
            system_message += "\n7. Respond in English"
        
        payload = {
            "model": "gpt-4o-mini",
            "messages": [
                {"role": "system", "content": system_message},
                {"role": "user", "content": user_message}
            ],
            "temperature": 0.7,
            "max_tokens": 200
        }
        
        response = requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers=headers,
            json=payload,
            timeout=15
        )
        
        if response.status_code == 200:
            ai_response = response.json()["choices"][0]["message"]["content"].strip()
            print(f"✅ AI Response: {ai_response[:80]}...")
            sys.stdout.flush()
            return ai_response, False, None
        else:
            print(f"❌ OpenAI Error: {response.status_code}")
            sys.stdout.flush()
            return "Sorry, I'm having trouble processing your request. Please try again or contact us directly! 📞", False, None
            
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        sys.stdout.flush()
        return "Sorry, I encountered an error. Please contact us directly! 📞", False, None

def send_typing_indicator(recipient_id):
    """Send typing indicator to show bot is processing"""
    try:
        url = f"https://graph.facebook.com/v21.0/{PAGE_ID}/messages"
        
        payload = {
            "recipient": {"id": recipient_id},
            "sender_action": "typing_on"
        }
        
        params = {"access_token": PAGE_ACCESS_TOKEN}
        
        response = requests.post(url, json=payload, params=params, timeout=5)
        return response.status_code == 200
    except:
        return False

def send_message_with_quick_replies(recipient_id, message_text, language):
    """Send message with quick reply buttons"""
    try:
        print(f"📤 Sending message with quick replies to {recipient_id}...")
        sys.stdout.flush()
        
        url = f"https://graph.facebook.com/v21.0/{PAGE_ID}/messages"
        
        quick_replies = get_quick_reply_buttons(language)
        
        payload = {
            "recipient": {"id": recipient_id},
            "message": {
                "text": message_text,
                "quick_replies": [
                    {
                        "content_type": "text",
                        "title": btn["title"],
                        "payload": btn["payload"]
                    }
                    for btn in quick_replies
                ]
            }
        }
        
        params = {"access_token": PAGE_ACCESS_TOKEN}
        
        response = requests.post(url, json=payload, params=params, timeout=10)
        
        if response.status_code == 200:
            print(f"✅ Message with quick replies sent!")
            sys.stdout.flush()
            return True
        else:
            print(f"❌ Error: {response.status_code}")
            sys.stdout.flush()
            return False
            
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        sys.stdout.flush()
        return False

def send_message(recipient_id, message_text):
    """Send message via Facebook Messenger API"""
    try:
        print(f"📤 Sending to {recipient_id}...")
        sys.stdout.flush()
        
        url = f"https://graph.facebook.com/v21.0/{PAGE_ID}/messages"
        
        payload = {
            "recipient": {"id": recipient_id},
            "messaging_type": "RESPONSE",
            "message": {"text": message_text}
        }
        
        params = {"access_token": PAGE_ACCESS_TOKEN}
        
        response = requests.post(url, json=payload, params=params, timeout=10)
        
        if response.status_code == 200:
            print(f"✅ Message sent successfully!")
            sys.stdout.flush()
            return True
        else:
            print(f"❌ Facebook Error: {response.status_code}")
            sys.stdout.flush()
            return False
            
    except Exception as e:
        print(f"❌ Error sending message: {str(e)}")
        sys.stdout.flush()
        return False

@app.route("/health", methods=["GET"])
def health():
    return {"status": "ok", "service": "Ultiphoton Chatbot"}, 200

@app.route("/", methods=["GET"])
def home():
    return "🤖 Ultiphoton Solar Power OPC Advanced Chatbot is running!", 200

@app.route("/privacy-policy", methods=["GET"])
def privacy_policy():
    html = """
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Privacy Policy - UltiPhoton Solar Power OPC</title>
  <style>
    body { font-family: Arial, sans-serif; max-width: 800px; margin: 40px auto; padding: 0 20px; color: #333; line-height: 1.7; }
    h1 { color: #f5a623; } h2 { color: #444; margin-top: 30px; }
    a { color: #f5a623; } footer { margin-top: 40px; font-size: 0.85em; color: #888; }
  </style>
</head>
<body>
  <h1>&#9728; UltiPhoton Solar Power OPC</h1>
  <h1>Privacy Policy</h1>
  <p><strong>Effective Date:</strong> April 8, 2026</p>
  <p>UltiPhoton Solar Power OPC ("we", "our", or "us") operates a Messenger chatbot to assist customers with inquiries about solar panels, accessories, and installation services. This Privacy Policy explains how we collect, use, and protect your information.</p>

  <h2>1. Information We Collect</h2>
  <p>When you interact with our Messenger chatbot, we may collect:</p>
  <ul>
    <li>Your Facebook name and profile ID (provided automatically by Facebook Messenger)</li>
    <li>Messages and inquiries you send to our chatbot</li>
    <li>Delivery details you voluntarily provide (name, address, contact number)</li>
  </ul>

  <h2>2. How We Use Your Information</h2>
  <p>We use the information collected to:</p>
  <ul>
    <li>Respond to your product and pricing inquiries</li>
    <li>Provide delivery fee estimates and arrange logistics</li>
    <li>Improve our chatbot responses and customer service</li>
    <li>Contact you regarding your order or inquiry</li>
  </ul>

  <h2>3. Data Sharing</h2>
  <p>We do not sell, trade, or share your personal information with third parties, except:</p>
  <ul>
    <li>With logistics partners (e.g., Lalamove) solely for delivery coordination</li>
    <li>As required by law or government authorities</li>
  </ul>

  <h2>4. Data Retention</h2>
  <p>Conversation data is stored temporarily to improve chatbot performance and is not retained beyond what is necessary for customer service purposes.</p>

  <h2>5. Your Rights</h2>
  <p>You may request deletion of your data at any time by contacting us directly through our Facebook Page: <a href="https://www.facebook.com/UltiPhotonSolarPowerOPC" target="_blank">UltiPhoton Solar Power OPC</a>.</p>

  <h2>6. Security</h2>
  <p>We implement reasonable security measures to protect your information from unauthorized access or disclosure.</p>

  <h2>7. Changes to This Policy</h2>
  <p>We may update this Privacy Policy from time to time. Changes will be posted on this page with an updated effective date.</p>

  <h2>8. Contact Us</h2>
  <p>If you have questions about this Privacy Policy, please contact us:</p>
  <ul>
    <li><strong>Business:</strong> UltiPhoton Solar Power OPC</li>
    <li><strong>Location:</strong> Brgy. Bolbok, Batangas City, Philippines</li>
    <li><strong>Facebook:</strong> <a href="https://www.facebook.com/UltiPhotonSolarPowerOPC" target="_blank">UltiPhoton Solar Power OPC</a></li>
  </ul>

  <footer><p>&copy; 2026 UltiPhoton Solar Power OPC. All rights reserved.</p></footer>
</body>
</html>
"""
    return html, 200, {"Content-Type": "text/html"}


@app.route("/webhook", methods=["GET", "POST"])
def webhook():
    if request.method == "GET":
        verify_token = request.args.get("verify_token")
        challenge = request.args.get("challenge")
        
        if verify_token == VERIFY_TOKEN:
            return challenge
        return "Invalid token", 403
    
    elif request.method == "POST":
        # ── Security: Verify X-Hub-Signature-256 (Meta HMAC) ──────────────────
        if APP_SECRET:
            sig_header = request.headers.get("X-Hub-Signature-256", "")
            if not sig_header.startswith("sha256="):
                print("⚠️  Webhook: missing X-Hub-Signature-256 header — rejecting request")
                sys.stdout.flush()
                return "Forbidden", 403
            expected_sig = "sha256=" + hmac.new(
                APP_SECRET.encode("utf-8"),
                request.get_data(),
                hashlib.sha256
            ).hexdigest()
            if not hmac.compare_digest(sig_header, expected_sig):
                print("⚠️  Webhook: invalid signature — rejecting request")
                sys.stdout.flush()
                return "Forbidden", 403
        # ──────────────────────────────────────────────────────────────────────
        try:
            data = request.get_json()
            
            if data.get("object") == "page":
                for entry in data.get("entry", []):
                    for messaging in entry.get("messaging", []):
                        sender_id = messaging.get("sender", {}).get("id")
                        message = messaging.get("message", {}).get("text")
                        
                        if sender_id and message:
                            print(f"\n📨 Message from {sender_id}: {message}")
                            sys.stdout.flush()
                            
                            # Detect language
                            detected_lang = detect_language(message)
                            saved_lang = get_user_language(sender_id)
                            
                            if saved_lang == "auto":
                                language = detected_lang
                                save_user_language(sender_id, language)
                            else:
                                language = saved_lang
                            
                            print(f"🌍 Language: {language}")
                            sys.stdout.flush()
                            
                            # Send auto-greeting if first message
                            if is_first_message(sender_id):
                                greeting = get_greeting_message(language)
                                send_message_with_quick_replies(sender_id, greeting, language)
                                mark_first_message_sent(sender_id)
                                time.sleep(0.3)
                            
                            # Send typing indicator
                            send_typing_indicator(sender_id)
                            time.sleep(0.3)
                            
                            # ── Step 1a: Check if user is replying with price tier (retail/installer) ──
                            pending_wattage = get_pending_wattage(sender_id)
                            pending_qty_for_tier = get_pending_panel_qty(sender_id)
                            if pending_wattage and pending_qty_for_tier:
                                tier_choice = detect_price_tier_reply(message)
                                if tier_choice:
                                    clear_pending_wattage(sender_id)
                                    clear_pending_panel_qty(sender_id)
                                    response_text = format_panel_package_response(pending_qty_for_tier, pending_wattage, language, price_tier=tier_choice)
                                    faq_matched, faq_key = True, "panel_package_calc"
                                    print(f"☀️ Panel package: {pending_qty_for_tier} pcs {pending_wattage}W ({tier_choice})") 
                                    sys.stdout.flush()
                                    log_analytics(sender_id, faq_key, message)
                                    save_conversation(sender_id, message, response_text, language, faq_matched)
                                    send_message_with_quick_replies(sender_id, response_text, language)
                                    return "EVENT_RECEIVED", 200
                                # Still waiting — fall through

                            # ── Step 1b: Check if user is replying with wattage (620/585) ──
                            pending_qty = get_pending_panel_qty(sender_id)
                            # Fallback: recover panel qty from conversation history if DB state was lost (e.g. after redeploy)
                            if not pending_qty and not pending_wattage:
                                pending_qty = get_recent_wattage_question(sender_id)
                                if pending_qty:
                                    print(f"🔄 Recovered pending_qty={pending_qty} from conversation history")
                                    sys.stdout.flush()
                            if pending_qty and not pending_wattage:
                                wattage_choice = detect_wattage_reply(message)
                                if wattage_choice:
                                    # 10+ panels: auto installer price; <10: retail price
                                    price_tier = 'installer' if pending_qty >= _INSTALLER_THRESHOLD else 'retail'
                                    clear_pending_panel_qty(sender_id)
                                    response_text = format_panel_package_response(pending_qty, wattage_choice, language, price_tier=price_tier)
                                    faq_matched, faq_key = True, "panel_package_calc"
                                    print(f"☀️ Panel package: {pending_qty} pcs {wattage_choice}W ({price_tier})")
                                    sys.stdout.flush()
                                    log_analytics(sender_id, faq_key, message)
                                    save_conversation(sender_id, message, response_text, language, faq_matched)
                                    send_message_with_quick_replies(sender_id, response_text, language)
                                    return "EVENT_RECEIVED", 200
                                # Still waiting — remind them to choose
                                # (fall through to normal AI response)

                            # ── Step 2: Detect panel qty inquiry without wattage ──
                            ambiguous_qty = detect_panel_qty_no_wattage(message)
                            if ambiguous_qty:
                                save_pending_panel_qty(sender_id, ambiguous_qty)
                                response_text = ask_wattage_question(ambiguous_qty, language)
                                faq_matched, faq_key = True, "wattage_clarification"
                                print(f"❓ Wattage clarification needed: {ambiguous_qty} panels")
                                sys.stdout.flush()
                                log_analytics(sender_id, faq_key, message)
                                save_conversation(sender_id, message, response_text, language, faq_matched)
                                send_message_with_quick_replies(sender_id, response_text, language)
                                return "EVENT_RECEIVED", 200

                            # Get AI response (with FAQ checking)
                            response_text, faq_matched, faq_key = get_ai_response(message, language)
                            
                            # Log analytics
                            if faq_matched and faq_key:
                                log_analytics(sender_id, faq_key, message)
                            
                            # Save conversation
                            save_conversation(sender_id, message, response_text, language, faq_matched)
                            
                            # Send response with quick replies
                            send_message_with_quick_replies(sender_id, response_text, language)
                            
            
            return "EVENT_RECEIVED", 200
            
        except Exception as e:
            print(f"❌ Webhook Error: {str(e)}")
            sys.stdout.flush()
            return "ERROR", 500

@app.route("/analytics", methods=["GET"])
def analytics():
    """Analytics dashboard endpoint — protected by ANALYTICS_TOKEN bearer auth"""
    # ── Security: require Bearer token when ANALYTICS_TOKEN env var is set ────────────
    if ANALYTICS_TOKEN:
        auth_header = request.headers.get("Authorization", "")
        token_param = request.args.get("token", "")
        provided = ""
        if auth_header.startswith("Bearer "):
            provided = auth_header[len("Bearer "):].strip()
        elif token_param:
            provided = token_param.strip()
        if not hmac.compare_digest(provided, ANALYTICS_TOKEN):
            return json.dumps({"error": "Unauthorized"}), 401, {"Content-Type": "application/json"}
    # ──────────────────────────────────────────────────────────────────────
    try:
        analytics_data = get_analytics_summary()
        return json.dumps(analytics_data, indent=2), 200, {"Content-Type": "application/json"}
    except Exception as e:
        return json.dumps({"error": str(e)}), 500, {"Content-Type": "application/json"}

if __name__ == "__main__":
    port = int(os.getenv("PORT", 10000))
    app.run(host="0.0.0.0", port=port, debug=False)
